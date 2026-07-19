# -*- coding: utf-8 -*-
r"""
build_index.py — 构建 GEO-ring Cloud 任务文件索引（sqlite + xlsx）

只读扫描 D:\AAAresearch_paper 下与 GEO-ring Cloud 任务相关的目录，
结合人工已核实的结构化元数据（相关性判定、脚本职责、外部路径引用、
流水线阶段、time_runs 批次），生成：
  - geo_ring_cloud_index.sqlite   （机器可读主存储，7 张表）
  - geo_ring_cloud_index.xlsx     （人机兼顾多 sheet 表格视图）

不修改任何已有文件，所有产出写入本脚本所在目录 _GEO_RING_CLOUD_INDEX/。
"""
from __future__ import annotations

import os
import sqlite3
import sys
import csv
import json
import re
from datetime import datetime, timezone
from pathlib import Path

import governance_check

ROOT = Path(r"D:\AAAresearch_paper")
OUT_DIR = Path(r"D:\AAAresearch_paper\_GEO_RING_CLOUD_INDEX")
DB_PATH = OUT_DIR / "geo_ring_cloud_index.sqlite"
REFRESHED_DB_PATH = OUT_DIR / "geo_ring_cloud_index_refreshed.sqlite"
XLSX_PATH = OUT_DIR / "geo_ring_cloud_index.xlsx"
WORKSPACE_DIR = ROOT / "_GEO_RING_CLOUD_WORKSPACE"
ARCHIVE_DIR = ROOT / "_NON_GEO_ARCHIVE"
GENERATED_AT = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

# ---------------------------------------------------------------------------
# 1. 目录扫描工具
# ---------------------------------------------------------------------------

def dir_stats(p: Path):
    """递归统计目录内文件数与总体积（字节）。跳过 __pycache__ / node_modules。"""
    nfiles = 0
    nbytes = 0
    exts = {}
    if not p.exists():
        return 0, 0, {}
    for dirpath, dirnames, filenames in os.walk(p):
        # 剪枝：跳过缓存/依赖目录，不计入统计
        dirnames[:] = [d for d in dirnames if d not in
                       ("__pycache__", "node_modules", ".git", ".claude", "dist", "venv", "_tmp")]
        for f in filenames:
            fp = Path(dirpath) / f
            try:
                sz = fp.stat().st_size
            except OSError:
                sz = 0
            nfiles += 1
            nbytes += sz
            ext = fp.suffix.lower() or "(无扩展名)"
            exts[ext] = exts.get(ext, 0) + 1
    return nfiles, nbytes, exts


def list_subdirs(p: Path, depth=1):
    """列出目录下指定深度的子目录。"""
    out = []
    if not p.exists():
        return out
    if depth == 1:
        for d in sorted(p.iterdir()):
            if d.is_dir() and not d.name.startswith(".") and d.name not in ("__pycache__", "node_modules"):
                out.append(d)
    return out


def fmt_size(b: int) -> str:
    if b is None:
        return ""
    for unit in ("B", "KB", "MB", "GB", "TB"):
        if b < 1024 or unit == "TB":
            return f"{b:.1f} {unit}" if unit != "B" else f"{b} B"
        b /= 1024
    return f"{b:.1f} TB"


def bool_text(value: bool) -> str:
    return "是" if value else "否"


def path_risk(relevance: str, referenced_by_code: str, exists_now: bool) -> str:
    if not exists_now:
        return "missing"
    if relevance == "强相关" or referenced_by_code in ("是", "间接"):
        return "keep_in_place"
    if relevance == "无关":
        return "archive_candidate"
    return "document_only"


def move_candidate(relevance: str, referenced_by_code: str, exists_now: bool) -> str:
    return bool_text(exists_now and relevance == "无关" and referenced_by_code == "否")


def infer_stage_from_name(rel_name: str) -> str:
    name = rel_name.replace("\\", "/")
    base = Path(name).name
    if base == "stage1_common.py":
        return "公共"
    if base.startswith("stage_10p2_") or name.startswith("stage_10p2_") or "/stage_10p2_" in name:
        return "10p2"
    if base.startswith("stage_10p_") or name.startswith("stage_10p_") or "/stage_10p_" in name:
        return "10p"
    if base.startswith("stage_10_") or name.startswith("stage_10_") or "/stage_10_" in name or "/stage_10_cth_validation/" in name:
        return "10"
    if base.startswith("stage_09f_") or name.startswith("stage_09f_") or "/stage_09f_" in name:
        return "09f"
    if base.startswith("stage_09e_") or name.startswith("stage_09e_") or "/stage_09e_" in name:
        return "09e"
    if base.startswith("stage_09d_") or name.startswith("stage_09d_") or "/stage_09d_" in name:
        return "09d"
    if base.startswith("stage09d_") or name.startswith("stage09d_") or "/stage09d_" in name:
        return "09d"
    if base.startswith("stage09c_") or name.startswith("stage09c_") or "/stage09c_" in name:
        return "09c"
    if base.startswith("stage09b_") or name.startswith("stage09b_") or "/stage09b_" in name:
        return "09b"
    prefix = base.split("_", 1)[0]
    if prefix[:2].isdigit():
        return prefix
    return ""


def summarize_script(rel_name: str) -> str:
    stage = infer_stage_from_name(rel_name)
    if stage:
        return f"当前文件系统扫描补充脚本；推断阶段 {stage}"
    return "当前文件系统扫描补充脚本"


# ---------------------------------------------------------------------------
# 2. 结构化元数据（已通过读码 + 子代理 + 亲自 Grep 核实）
# ---------------------------------------------------------------------------

# 顶层目录相关性判定
# relevance: 强相关 / 上游相关 / 弱相关 / 无关
# referenced_by_code: 是否被 ring cloud 代码硬编码引用
TOP_DIRS = [
    # (相对路径, 相关性, 角色, 体积说明, 被代码引用, 备注)
    ("geo_ring_cloud_stage1", "强相关", "Stage1 主产物根目录（STAGE_ROOT）", "约 6.4 GB", "是", "standardized_native/reprojected_grid/fused_best_source/overlap_validation/reports 等产物 + scripts 副本"),
    ("geo_ring_cloud_stage1_time_runs", "强相关", "多时次运行、跨阶段实验与阶段诊断产物根目录", "较大", "是", "运行器与 stage_08 以后诊断脚本通过 RUNS_ROOT 引用；大产物保留原位"),
    ("geo_ring_cloud_stage1_evidence_pack", "强相关", "Stage1 证据包（latest + 10 snapshots）", "约 1.2 MB", "是", "rebuild_stage1_evidence_pack.py:EVIDENCE_ROOT 产出"),
    ("third_report/code/geo_ring_cloud_stage1", "强相关", "GEO-ring Cloud 主代码、阶段脚本、共享组件与测试", "代码目录", "是", "权威代码源；产物写入配置化的 stage/time-runs 根目录"),
    ("data_check_report", "强相关", "前序数据审计报告（REPORT_ROOT，00-00f 阶段证据源）", "约 872 MB / 185 文件", "是", "geo_ring_cloud.paths + pipeline_layout 中的 REPORT_ROOT、PARSED_METADATA、MAPPING_YAML"),
    ("geo_geometry_check", "强相关", "几何校验样本（download_geo_geometry_samples.py 产物 + 06c/06d 审计）", "约 1.34 GB / 50 文件", "是", "download_geo_geometry_samples.py:22 OUT_ROOT；06c/06d/06e 引用"),
    ("data", "强相关", "原始卫星数据（FY4A/FY4B L1 + FY4B 云产品 CLM/CLP/CLT/CTH/CTP/CTT/GEO + H09_Data）", "约 100+ GB", "是", "geo_ring_cloud.paths.HIMAWARI_R21_DIR；06f 扫描 data/"),
    ("third_report/code/L1g", "上游相关", "standardized_L1_source 规范 + 全球 0.05° 网格规范", "小", "间接", "ring cloud 上游标准化层规范文档"),
    ("third_report/code/FY4B", "上游相关", "FY4B/AGRI standardized_L1 builder + 探查/预览", "小", "间接", "产出 standardized_L1_source 供 stage1 输入"),
    ("third_report/code/GOES", "上游相关", "GOES-16/18/ABI standardized_L1 builder", "小", "间接", "同上"),
    ("third_report/code/Himawari", "上游相关", "Himawari-9/AHI standardized_L1 builder", "小", "间接", "同上"),
    ("third_report/code/Meteosat", "上游相关", "Meteosat-9/10/SEVIRI standardized_L1 builder", "小", "间接", "同上"),
    ("third_report/code/geo_data_audit", "上游相关", "数据审计脚本（前序审计 00-00f 的执行代码）", "小", "间接", "audit_geometry_and_variables.py 等产出 data_check_report"),
    ("third_report/code/geo_cloud_download", "上游相关", "GEO 云产品下载器（EUMETSAT API + S3）", "小", "间接", "下载原始云产品到 data/ 与 E 盘"),
    ("third_report/code/priority_download_goes_meteosat", "上游相关", "GOES/Meteosat 优先补下载（00f 阶段）", "小", "间接", "build/download/verify 三件套"),
    ("third_report/code/preview_baselines", "上游相关", "基线预览图索引", "小", "间接", "六星基线 quicklook"),
    ("third_report/code/standardized_l1_source_satpy.py", "上游相关", "Satpy 共用标准化骨架", "小", "间接", "被 Himawari/Meteosat builder 调用"),
    ("third_report/code/run_standardized_l1_source_batch.py", "上游相关", "标准化批处理统一入口", "小", "间接", "调度各卫星 builder"),
    ("third_report/code/validate_standardized_l1_source_samples.py", "上游相关", "标准化样例自动校验", "小", "间接", ""),
    ("third_report/code/preview_runner.py", "上游相关", "预览批处理入口", "小", "间接", ""),
    ("third_report/code/plot_geo_satellite_coverage.py", "上游相关", "六星地球覆盖边界图", "小", "间接", ""),
    ("third_report/Satellite_Data_20240312", "上游相关", "2024-03-12 主数据集快照（六星原生 + standardized_L1 样例 + reprojected 旧版）", "大", "间接", "含 channel_mapping/standardized_L1_source 等"),
    ("research_tracker", "弱相关", "研究追踪器（项目治理元工具，扫描全项目建知识图谱）", "中", "否", "元工具，不参与流水线；识别 ring_cloud 为最大子项目"),
    ("cloud", "弱相关", "云产品参考文献 PDF（Zhao 等 2026 全天全球云物理属性）", "约 1.2 MB", "否", "方法论参考，代码无路径引用"),
    ("third_report/beamer", "弱相关", "LaTeX Beamer 组会幻灯片（讲标准化上游）", "中", "否", "讲稿明确讲 L1g 前期工作，非 ring cloud 本体"),
    ("6月第二次汇报邓浩然.pptx", "弱相关", "6 月组会汇报 PPT", "约 52 MB", "否", "项目阶段汇报材料"),
    ("EPIC数据.pptx", "弱相关", "EPIC 数据 PPT", "约 38 KB", "否", "偏向 EPIC/CERES 分支"),
    ("forth", "无关", "导航相机 DAT 数据解析 + QuickView 工具（独立任务）", "约 1.3 GB", "否", "nav_camera_dat.py；ring cloud 代码零引用"),
    ("second_report", "无关", "FY 数据分析第二次报告（FYDataService + 探查 notebook + 论文）", "约 18 GB", "否", "ring cloud 代码零引用"),
    ("third_report/code/epic_ceres", "无关", "EPIC→CERES 深度学习独立分支（Stage 1-13）", "中", "否", "估算 CERES TOA 辐射通量，与云产品融合无关；代码零引用"),
    ("third_report/outputs/epic_ceres", "无关", "EPIC-CERES 各阶段输出", "大", "否", "epic_ceres 分支产物"),
    ("third_report/outputs/epic_ceres_fullmonth_v2", "无关", "EPIC-CERES 全月 v2 训练输出", "大", "否", "epic_ceres 分支产物"),
    ("third_report/outputs/epic_ceres_ppt_assets", "无关", "EPIC-CERES PPT 素材", "小", "否", "epic_ceres 分支产物"),
    ("third_report/paper_notion_manager", "无关", "论文-Notion 管理工具", "小", "否", "LLM+Notion 管理论文，与气象无关"),
    ("data/3-科大蓝（竞赛 科技）.pptx", "无关", "存于 data/ 下的竞赛演示文稿", "约 43 MB", "否", "与 ring cloud 无关"),
    ("data/A202607010902106154", "无关", "空目录", "0", "否", "空"),
]

# 外部数据盘（不在 D:\AAAresearch_paper 内）
EXTERNAL_DISKS = [
    ("E:\\GEO_Cloud_2024", "E盘", "GOES/Himawari/Meteosat 大规模原始云产品存储根（check_download_completeness.py GEO_ROOT；08b 引用 EPIC L2 CMSAF 子目录）", "是"),
    ("F:\\DSCOVR_EPIC_L2_CLOUD_03_2024.03", "F盘", "DSCOVR EPIC L2 云产品数据（time_runs manifest 中 single_sample_run_manifest.json 引用）", "是"),
]

# ---------------------------------------------------------------------------
# 3. ring cloud 主脚本职责表（third_report/code/geo_ring_cloud_stage1/ 下 41 个 .py）
#    (文件名, 阶段, 职责)
# ---------------------------------------------------------------------------
SCRIPTS = [
    ("path_config.py", "common", "Central environment-overridable paths including GEO_RING_CLAAS3_ROOT"),
    ("geo_ring_cloud_source_registry.py", "common", "Stable source IDs, processing streams, products, profiles, tolerances, and variable rules"),
    ("geo_ring_cloud_claas3_adapter.py", "common", "CLAAS-3 recursive discovery, deterministic deduplication, decoding, unit conversion, QA, and per-variable masks"),
    ("geo_ring_cloud_lineage.py", "common", "Common run and artifact lineage manifest writer"),
    ("geo_ring_cloud_run_discovery.py", "common", "Matrix-manifest-first run discovery with legacy time-tag compatibility"),
    ("geo_ring_cloud_time_run_matrix.py", "runner", "Matched operational_baseline and claas3_candidate runner with SQLite indexing"),
    ("stage_00d_claas3_integration_readiness.py", "00d", "CLAAS-3 March 2024 cadence, structure, QA, projection, and integration readiness gate"),
    ("stage_06c_claas3_geometry_angle_lineage.py", "06c", "CLAAS-3 CF projection and navigation-derived angle lineage gate"),
    ("stage_07p_claas3_profile_pair_evaluation.py", "07p", "Common-domain CLAAS-3 versus operational Meteosat consistency and boundary diagnostics"),
    ("stage_09d_claas3_epic_profile_pair_evaluation.py", "09d", "Matched common-domain EPIC cloud-mask profile-pair metrics and sample-block bootstrap"),
    ("stage_10_claas3_epic_relative_height_evaluation.py", "10", "A/B-band EPIC-relative effective-height profile-pair diagnostics with common approximate PSF"),
    ("stage1_common.py", "公共", "已登记 compatibility shim；权威 API 见 geo_ring_cloud.pipeline_support、pipeline_layout、cloud_semantics 和 diagnostics.summary"),
    ("01_build_core_time_index.py", "01", "从 data_check_report/parsed_file_metadata.csv 构建核心时次索引，按卫星完整度评分，选定原型时次 2024-03-05T00:00Z"),
    ("02_build_standardized_cloud_native.py", "02", "读取各卫星原生云产品，按统一变量名映射标准化为 native-grid NPZ，输出到 standardized_native/"),
    ("03_validate_standardized_cloud_native.py", "03", "校验 02 产物 NPZ 的变量完整性、dtype、shape 一致性"),
    ("03_5_semantic_validation_patch.py", "03.5", "语义校验补丁：分类变量码值范围、fill 值合理性"),
    ("04_check_fy4b_geo_alignment.py", "04", "FY4B GEO 产品与 L2 云产品网格对齐检查（shape/变量/角度范围）"),
    ("04b_fy4b_dqf_bit_decode_diagnostics.py", "04b", "FY4B DQF 质量标志位级解码诊断，生成码值统计"),
    ("05_reproject_cloud_to_grid.py", "05", "各卫星原生网格重投影到统一 0.05° 全球网格(3600x7200)，KD-tree 最近邻；输出 display/fusion valid_mask"),
    ("06_fuse_best_source.py", "06", "变量级 best-source 融合：基于 VZA/view_weight/time_weight 逐像素选最优源；输出融合数据+source_map+rating_map"),
    ("06_5_source_selection_diagnostics.py", "06.5", "源选择诊断：验证融合是否以 min-VZA 逻辑驱动"),
    ("06c_geometry_parameter_audit.py", "06c", "几何参数审计：提取各卫星子午经度/地球半径/轨道高度等"),
    ("06c_multi_satellite_geometry_metadata_audit.py", "06c", "多卫星几何元数据审计（引用 geo_geometry_check + reprojected_grid + standardized_native）"),
    ("06d_himawari_full_disk_geometry_validation.py", "06d", "Himawari 全圆盘几何验证（引用 geo_geometry_check/Himawari-9 与 vza_method_comparison_by_satellite.csv）"),
    ("06e_full_geometry_angle_source_sync_patch.py", "06e", "几何角度源同步补丁：将传感器/太阳角度层投影到目标网格并重跑 06 融合"),
    ("06e_vza_ecef_final_audit.py", "06e", "VZA ECEF 坐标系最终审计（引用 geo_geometry_check）"),
    ("06f_unknown_aware_data_asset_audit.py", "06f", "unknown-aware 数据资产审计：扫描 data/geo_geometry_check/stage1 子目录识别未知变量"),
    ("06f_reexport_with_obitype_patch.py", "06f", "带 orbit type 补丁的重新导出"),
    ("06f_report_sync_patch.py", "06f", "报告同步补丁"),
    ("07_overlap_consistency_validation.py", "07", "重叠区一致性验证 v1：相邻卫星覆盖区 cloud_mask/CTH/CTT 差异（历史版）"),
    ("07p_overlap_validator_hotfix.py", "07p", "重叠验证热修复：修 cloud-mask 映射/angle-layer/分层执行"),
    ("07p_b_source_boundary_magnitude_review.py", "07p-b", "源边界跳变幅度审查"),
    ("07v2_formal_single_time_report.py", "07v2", "正式单时次报告生成：聚合 07p，生成最终验收决策"),
    ("08_epic_visual_comparison.py", "08", "EPIC(DSCOVR) 目视比较：下载 EPIC 图像与融合结果对比"),
    ("08b_epic_l2_cloud_audit_compare.py", "08b", "EPIC L2 云产品审计比较（引用 E:\\GEO_Cloud_2024\\CMSAF 下 EPIC L2 文件）"),
    ("08c_epic_cloud_mask_semantic_sensitivity.py", "08c", "EPIC 云掩膜语义敏感性分析"),
    ("08d_select_epic_monthly_semantic_validation_targets.py", "08d", "选择 EPIC 月度语义验证目标"),
    ("08e_summarize_epic_georing_multisample.py", "08e", "多样本 EPIC Geo-ring 汇总（RUNS_ROOT=time_runs）"),
    ("08f_geometry_and_prefusion_epic_diagnostics.py", "08f", "几何与预融合 EPIC 诊断（RUNS_ROOT=time_runs）"),
    ("08g_overlap_count_diagnostics.py", "08g", "重叠计数诊断（RUNS_ROOT=time_runs）"),
    ("08h_meteosat_time_offset_control_test.py", "08h", "Meteosat 时间偏移控制测试（RUNS_ROOT=time_runs）"),
    ("08i_meteosat_source_overlap_diagnostics.py", "08i", "Meteosat 源重叠诊断（RUNS_ROOT=time_runs）"),
    ("08j_prefusion_source_pair_overlap_diagnostics.py", "08j", "预融合源对重叠诊断（RUNS_ROOT=time_runs）"),
    ("08k_consolidate_stage08_report.py", "08k", "Stage08 报告整合（RUNS_ROOT=time_runs）"),
    ("09_stage09_epic_georing_cloud_mask_diagnostics.py", "09", "Stage09 EPIC-Geo-ring 云掩膜诊断（RUNS_ROOT=time_runs）"),
    ("stage09b_full_overnight/run_stage09b_full_overnight.py", "09b", "Stage09b 全夜批量运行（RUNS_ROOT + BASE_STAGE_ROOT）"),
    ("stage09c_scaled_batch/run_stage09c_scaled_batch.py", "09c", "Stage09c 扩展批量运行（RUNS_ROOT + BASE_STAGE_ROOT）"),
    ("download_geo_geometry_samples.py", "下载", "从 AWS S3 下载 GOES-16/18/Himawari-9 几何样本到 geo_geometry_check/"),
    ("rebuild_stage1_evidence_pack.py", "证据包", "重建 Stage1 证据包：汇总 data_check_report/geo_geometry_check/stage1 全部证据到 evidence_pack/"),
    ("run_epic_georing_single_sample.py", "运行器", "EPIC Geo-ring 单时次完整运行流水线（BASE=stage1, RUNS=time_runs）"),
    ("run_epic_georing_sample_batch.py", "运行器", "EPIC Geo-ring 批量样本运行器（RUNS=time_runs）"),
    ("summarize_time_run_20240319_1500.py", "汇总", "汇总 20240319_1500 时次运行结果"),
]

# ---------------------------------------------------------------------------
# 4. 外部数据路径引用表（亲自 Grep 核实，含行号）
#    (外部路径, 引用脚本, 行号, 用途, 位置)
# ---------------------------------------------------------------------------
EXT_REFS = [
    # canonical paths and pipeline layout
    (r"D:\AAAresearch_paper\geo_ring_cloud_stage1", "geo_ring_cloud/paths.py", 20, "STAGE_ROOT 产物根（可被 GEO_RING_STAGE_ROOT 环境变量覆盖）", "D盘内"),
    (r"D:\AAAresearch_paper\data_check_report", "geo_ring_cloud/paths.py", 25, "DATA_CHECK_ROOT 前序审计报告根", "D盘内"),
    (r"D:\AAAresearch_paper\data_check_report\geometry_variable_audit", "geo_ring_cloud/paths.py", 26, "DATA_CHECK_GEOMETRY_ROOT 几何审计目录", "D盘内"),
    (r"D:\AAAresearch_paper\data\H09_Data", "geo_ring_cloud/paths.py", 32, "HIMAWARI_R21_DIR Himawari R21 辅助几何文件", "D盘内"),
    (r"D:\AAAresearch_paper\data_check_report\parsed_file_metadata.csv", "geo_ring_cloud/pipeline_layout.py", 14, "PARSED_METADATA 所有卫星文件元数据索引（01 阶段核心输入）", "D盘内"),
    (r"D:\AAAresearch_paper\data_check_report\geometry_variable_audit\product_variable_inventory_full.csv", "geo_ring_cloud/pipeline_layout.py", 15, "VARIABLE_INVENTORY 产品变量清单", "D盘内"),
    (r"D:\AAAresearch_paper\data_check_report\manual_variable_mapping_by_product.yaml", "geo_ring_cloud/pipeline_layout.py", 16, "MAPPING_YAML 手动变量映射表（02 直接读取）", "D盘内"),
    # download_geo_geometry_samples.py
    (r"D:\AAAresearch_paper\geo_geometry_check", "download_geo_geometry_samples.py", 22, "OUT_ROOT 几何样本下载产物根", "D盘内"),
    # 06c/06d/06e 对 geo_geometry_check 的引用
    (r"D:\AAAresearch_paper\geo_geometry_check", "06c_multi_satellite_geometry_metadata_audit.py", 27, "GEOMETRY_ROOT 几何样本根", "D盘内"),
    (r"D:\AAAresearch_paper\geo_geometry_check\vza_method_comparison_by_satellite.csv", "06d_himawari_full_disk_geometry_validation.py", 28, "CURRENT06C_VZA_CSV 当前 06c VZA 比较结果", "D盘内"),
    (r"D:\AAAresearch_paper\geo_geometry_check\Himawari-9", "06d_himawari_full_disk_geometry_validation.py", 29, "GEOMETRY_ROOT Himawari 全圆盘段数据", "D盘内"),
    (r"D:\AAAresearch_paper\geo_geometry_check", "06e_vza_ecef_final_audit.py", 20, "EXTERNAL_GEOMETRY_AUDIT_DIR 几何审计目录", "D盘内"),
    # 06f 扫描的多个数据目录
    (r"D:\AAAresearch_paper\data", "06f_unknown_aware_data_asset_audit.py", 38, "SCAN_DIRS 原始卫星数据根", "D盘内"),
    (r"D:\AAAresearch_paper\geo_geometry_check", "06f_unknown_aware_data_asset_audit.py", 39, "SCAN_DIRS 几何校验样本", "D盘内"),
    (r"D:\AAAresearch_paper\geo_ring_cloud_stage1\standardized_native", "06f_unknown_aware_data_asset_audit.py", 40, "SCAN_DIRS 标准化原生产物", "D盘内"),
    (r"D:\AAAresearch_paper\geo_ring_cloud_stage1\reprojected_grid", "06f_unknown_aware_data_asset_audit.py", 41, "SCAN_DIRS 重投影产物", "D盘内"),
    (r"D:\AAAresearch_paper\geo_ring_cloud_stage1\fused_best_source", "06f_unknown_aware_data_asset_audit.py", 42, "SCAN_DIRS 融合产物", "D盘内"),
    # rebuild_stage1_evidence_pack.py 汇总的所有证据源
    (r"D:\AAAresearch_paper\geo_ring_cloud_stage1_evidence_pack", "rebuild_stage1_evidence_pack.py", 14, "EVIDENCE_ROOT 证据包输出根", "D盘内"),
    (r"D:\AAAresearch_paper\geo_ring_cloud_stage1", "rebuild_stage1_evidence_pack.py", 15, "STAGE1_ROOT 主产物根（汇总证据来源）", "D盘内"),
    (r"D:\AAAresearch_paper\geo_geometry_check", "rebuild_stage1_evidence_pack.py", 16, "GEOMETRY_ROOT 几何证据来源", "D盘内"),
    (r"D:\AAAresearch_paper\third_report\code\geo_ring_cloud_stage1", "rebuild_stage1_evidence_pack.py", 17, "CODE_ROOT 主代码来源", "D盘内"),
    (r"D:\AAAresearch_paper\data_check_report", "rebuild_stage1_evidence_pack.py", 18, "DATA_CHECK_ROOT 前序审计证据来源", "D盘内"),
    # time_runs 引用
    (r"D:\AAAresearch_paper\geo_ring_cloud_stage1_time_runs", "08e_summarize_epic_georing_multisample.py", 16, "RUNS_ROOT 多时次运行根", "D盘内"),
    (r"D:\AAAresearch_paper\geo_ring_cloud_stage1_time_runs", "08f_geometry_and_prefusion_epic_diagnostics.py", 14, "RUNS_ROOT 多时次运行根", "D盘内"),
    (r"D:\AAAresearch_paper\geo_ring_cloud_stage1_time_runs", "08g_overlap_count_diagnostics.py", 13, "ROOT 多时次运行根", "D盘内"),
    (r"D:\AAAresearch_paper\geo_ring_cloud_stage1_time_runs", "08h_meteosat_time_offset_control_test.py", 15, "RUNS_ROOT 多时次运行根", "D盘内"),
    (r"D:\AAAresearch_paper\geo_ring_cloud_stage1_time_runs", "08i_meteosat_source_overlap_diagnostics.py", 14, "RUNS_ROOT 多时次运行根", "D盘内"),
    (r"D:\AAAresearch_paper\geo_ring_cloud_stage1_time_runs", "08j_prefusion_source_pair_overlap_diagnostics.py", 13, "RUNS_ROOT 多时次运行根", "D盘内"),
    (r"D:\AAAresearch_paper\geo_ring_cloud_stage1_time_runs", "08k_consolidate_stage08_report.py", 11, "RUNS_ROOT 多时次运行根", "D盘内"),
    (r"D:\AAAresearch_paper\geo_ring_cloud_stage1_time_runs", "09_stage09_epic_georing_cloud_mask_diagnostics.py", 21, "RUNS_ROOT 多时次运行根", "D盘内"),
    (r"D:\AAAresearch_paper\geo_ring_cloud_stage1_time_runs\20240319_1500", "08b_epic_l2_cloud_audit_compare.py", 24, "TIME_RUN_ROOT 特定时次运行目录", "D盘内"),
    (r"D:\AAAresearch_paper\geo_ring_cloud_stage1_time_runs\epic_202403_target_selection\recommended_epic_georing_validation_targets.csv", "run_epic_georing_sample_batch.py", 15, "DEFAULT_SELECTION_CSV EPIC 验证目标选择清单", "D盘内"),
    # 外部盘 E 盘
    (r"E:\GEO_Cloud_2024\CMSAF\DSCOVR_EPIC_L2_CLOUD_03_20240319150052_03.nc4", "08b_epic_l2_cloud_audit_compare.py", 27, "EPIC_L2_FILE DSCOVR EPIC L2 云产品文件", "E盘"),
    # 外部盘 F 盘（time_runs manifest 引用）
    (r"F:\DSCOVR_EPIC_L2_CLOUD_03_2024.03", "time_runs/*/single_sample_run_manifest.json", 5, "DSCOVR EPIC L2 云产品数据根（各批次 manifest 引用具体 .nc4 文件）", "F盘"),
    # check_download_completeness.py（在 data_check_report，引用 data 与 E 盘）
    (r"D:\AAAresearch_paper\data", "data_check_report/check_download_completeness.py", 17, "DATA_ROOT 原始卫星数据根", "D盘内"),
    (r"E:\GEO_Cloud_2024", "data_check_report/check_download_completeness.py", 18, "GEO_ROOT 外部数据盘 GEO 云产品存储根", "E盘"),
]

# ---------------------------------------------------------------------------
# 5. 流水线阶段表（源自 evidence_sources_by_stage.md + stage_registry.md）
#    (阶段, 名称, 输入, 输出, Gate状态, 证据源目录)
# ---------------------------------------------------------------------------
PIPELINE_STAGES = [
    ("00", "数据下载与完整性审计", "原始卫星数据下载", "data_download_audit_report.md, *_cross_product_match.csv", "Complete", "data_check_report/"),
    ("00b", "一样本一产品真实读取", "样本文件", "manual_variable_mapping_by_product.yaml, one_sample_each_product_*.csv", "Complete", "data_check_report/"),
    ("00c", "几何与变量能力审计", "样本文件", "product_variable_inventory_full.csv, geometry/cloud/quality_flag_capability_matrix.csv", "Complete", "data_check_report/geometry_variable_audit/"),
    ("00d", "Meteosat 产品体系与 catalogue discovery", "Meteosat 目录", "meteosat_product_series_audit.md, meteosat_catalogue_discovery_report.md", "Complete", "data_check_report/meteosat_*/"),
    ("00e", "单时次 v0 标准化原型", "样本", "standardized_cloud_v0_report.md, FY4B_CLM/GOES-16_ACMF/Himawari-9_CMSK.npz", "Complete", "data_check_report/standardized_cloud_v0_samples/"),
    ("00f", "GOES/Meteosat 优先补下载", "缺口清单", "priority_download_goes_meteosat_report.md, meteosat_priority_download_report.md", "Complete", "data_check_report/priority_download_run_*/"),
    ("01", "核心时间索引", "parsed_file_metadata.csv", "core_time_index.csv, usable_times_ranked.csv（原型 2024-03-05T00:00Z）", "PASS", "geo_ring_cloud_stage1/time_index/"),
    ("02", "标准化云原生 NPZ", "01 选定时次 + 原始卫星文件", "standardized_native/*.npz, inventory/stats CSV, quicklook", "PASS", "geo_ring_cloud_stage1/standardized_native/"),
    ("03", "结构校验", "02 的 NPZ", "standardized_native_file_validation.csv, validate report", "PASS", "geo_ring_cloud_stage1/standardized_native/"),
    ("03.5", "语义校验补丁", "02 的 NPZ", "standardized_native_semantic_issues.csv, code_tables.csv", "PASS_WITH_WARNINGS", "geo_ring_cloud_stage1/standardized_native/"),
    ("04", "FY4B GEO 对齐检查", "FY4B NPZ(CLM-GEO)", "fy4b_geo_alignment_*.csv, fy4b_geo_alignment_report.md", "PASS_WITH_WARNINGS", "geo_ring_cloud_stage1/reports/"),
    ("04b", "FY4B DQF 位解码诊断", "FY4B NPZ DQF", "fy4b_dqf_bit_decode_summary.csv, fy4b_quality_flag_rules.yaml", "PASS_WITH_WARNINGS", "geo_ring_cloud_stage1/reports/"),
    ("05", "重投影到 0.05° 网格", "02 NPZ + 04 GEO 对齐", "reprojected_grid/{6卫星}/*.npz, target_grid_definition.json, coverage map", "PASS_WITH_WARNINGS", "geo_ring_cloud_stage1/reprojected_grid/"),
    ("06", "best-source 融合", "05 重投影 NPZ", "fused_geo_ring_cloud_*.npz, source_map/rating_map, fusion_stats.csv", "PASS_WITH_WARNINGS", "geo_ring_cloud_stage1/fused_best_source/"),
    ("06.5", "源选择诊断", "06 融合结果", "selected_vs_min_vza_agreement.csv 等 9 个 CSV", "PASS", "geo_ring_cloud_stage1/source_selection_diagnostics/"),
    ("06c", "多星几何元数据审计", "05 重投影 + geo_geometry_check", "satellite_geometry_parameter_audit.csv, vza_method_comparison.csv", "PASS_WITH_WARNINGS", "geo_geometry_check/ + geometry_audit_06c/"),
    ("06d", "Himawari 全圆盘几何验证", "geo_geometry_check/Himawari-9", "himawari_full_disk_geometry_report.md/audit.csv", "PASS_WITH_WARNINGS", "geo_geometry_check/Himawari-9/"),
    ("06e", "几何角度源同步", "05/06 结果 + 角度层", "angle_provenance_inventory.csv, geometry_angle_source_policy.yaml, 角度层 NPZ", "PASS", "geo_ring_cloud_stage1/geometry_angle_sync_06e/"),
    ("06f", "unknown-aware 数据资产审计", "全部 NPZ 产物", "audit_summary.json, data_asset_audit.sqlite, blocking_issues.csv", "PASS_WITH_WARNINGS", "geo_ring_cloud_stage1/data_asset_audit_06f/"),
    ("07", "重叠一致性验证(原版)", "06 融合结果", "overlap_*.csv, overlap_confusion_matrices.json", "Historical", "geo_ring_cloud_stage1/overlap_validation/"),
    ("07p", "重叠验证修复版", "06e 修正后结果", "*_v2.csv, overlap_validation_07p_report.md", "PASS", "geo_ring_cloud_stage1/overlap_validation_07p/"),
    ("07p-b", "边界跳变幅度审查", "07p 结果", "07p_boundary_magnitude_review.md", "PASS", "geo_ring_cloud_stage1/reports/"),
    ("07v2", "正式单时次报告", "07p 全部结果", "07v2_overlap_consistency_validation_report.md, stage1_single_time_acceptance_decision.md", "PASS_WITH_WARNINGS", "geo_ring_cloud_stage1/reports/"),
    ("08", "EPIC 目视对比(原型时次)", "06 融合 + EPIC 图像", "epic_visual_comparison_report.md, epic_*.csv, 对比图", "—", "geo_ring_cloud_stage1/epic_visual_comparison/"),
    ("08c", "EPIC L2 云掩膜语义敏感性", "time_runs 各批次 + EPIC L2", "epic_l2_cloud_mask_semantic_sensitivity_*/", "—", "geo_ring_cloud_stage1_time_runs/*/"),
    ("09", "Stage09 EPIC-Geo-ring 诊断", "time_runs + EPIC", "stage09 系列诊断", "—", "geo_ring_cloud_stage1_time_runs/epic_202403_*/"),
]

# ---------------------------------------------------------------------------
# 6. time_runs 批次表（48 个时间批次 + 5 个 EPIC 专题，Glob 核实）
# ---------------------------------------------------------------------------
TIME_RUNS = [
    ("20240305_1500",), ("20240306_1100",), ("20240306_1300",), ("20240307_0900",),
    ("20240308_0400",), ("20240308_1300",), ("20240309_0100",), ("20240309_0800",),
    ("20240309_1300",), ("20240309_1700",), ("20240310_1000",), ("20240310_1200",),
    ("20240311_0800",), ("20240311_1000",), ("20240311_1400",), ("20240311_1900",),
    ("20240311_2300",), ("20240312_1500",), ("20240313_0400",), ("20240313_1100",),
    ("20240313_2200",), ("20240314_0900",), ("20240315_0400",), ("20240315_1300",),
    ("20240316_0800",), ("20240316_1700",), ("20240317_1000",), ("20240317_1200",),
    ("20240318_0800",), ("20240318_1000",), ("20240319_1500",), ("20240320_0400",),
    ("20240321_0900",), ("20240321_1100",), ("20240322_0400",), ("20240322_1300",),
    ("20240323_0800",), ("20240324_1200",), ("20240325_0800",), ("20240325_1000",),
    ("20240326_1500",), ("20240327_1100",), ("20240328_0900",), ("20240328_1100",),
    ("20240329_1300",), ("20240330_0800",), ("20240330_1700",), ("20240331_1200",),
    ("20240331_1900",),
]
EPIC_TOPICS = [
    ("epic_202403_batch_runs", "EPIC Geo-ring 批量运行状态与报告"),
    ("epic_202403_meteosat_time_offset_control", "Meteosat 时间偏移控制实验"),
    ("epic_202403_multisample_summary", "多样本 EPIC-Geo-ring 汇总 + Stage08 组会 PPT"),
    ("epic_202403_overnight_watch", "过夜观测"),
    ("epic_202403_target_selection", "EPIC 验证目标选择候选清单"),
]

PROJECT_ID = "geo_ring_cloud"
KEY_ARTIFACT_EXTS = {".md", ".csv", ".json", ".yaml", ".yml", ".toml", ".xlsx", ".py", ".ps1"}
AGGREGATE_EXTS = {".png", ".npz", ".nc", ".nc4", ".h5", ".hdf", ""}
COMPONENT_ROLES = {
    "common": "shared_library",
    "runner": "runner",
    "公共": "shared_library",
    "运行器": "runner",
    "下载": "downloader",
    "证据包": "evidence_pack_builder",
    "汇总": "summary_helper",
    "data_product_audit": "data_product_audit",
    "": "support",
}
MODULE_REGISTRY = (
    {
        "project_id": PROJECT_ID,
        "canonical_module": "geo_ring_cloud.paths",
        "canonical_path": "third_report/code/geo_ring_cloud_stage1/geo_ring_cloud/paths.py",
        "component_role": "path_configuration",
        "legacy_module": "path_config",
        "legacy_path": "third_report/code/geo_ring_cloud_stage1/path_config.py",
        "migration_status": "canonical_with_compatibility_shim",
        "public_api": "env_path and project/data/output/credential root constants",
        "test_evidence": "tests/geo_ring_cloud_test_claas3.py::PackageBoundaryTests",
        "notes": "Canonical Python location allowed to contain default machine-local paths.",
    },
    {
        "project_id": PROJECT_ID,
        "canonical_module": "geo_ring_cloud.path_configuration_powershell",
        "canonical_path": "third_report/code/geo_ring_cloud_stage1/geo_ring_cloud_path_configuration.ps1",
        "component_role": "path_configuration",
        "legacy_module": "",
        "legacy_path": "",
        "migration_status": "canonical",
        "public_api": "GeoRing project/data/output/credential/tool path variables for PowerShell orchestration",
        "test_evidence": "_GEO_RING_CLOUD_INDEX/tests/test_governance_check.py::PathEnforcementTests",
        "notes": "PowerShell companion to geo_ring_cloud.paths; uses the same GEO_RING_* environment-variable contract.",
    },
    {
        "project_id": PROJECT_ID,
        "canonical_module": "geo_ring_cloud.sources",
        "canonical_path": "third_report/code/geo_ring_cloud_stage1/geo_ring_cloud/sources.py",
        "component_role": "source_registry",
        "legacy_module": "geo_ring_cloud_source_registry",
        "legacy_path": "third_report/code/geo_ring_cloud_stage1/geo_ring_cloud_source_registry.py",
        "migration_status": "canonical_with_compatibility_shim",
        "public_api": "source definitions, IDs, source profiles, variable rules",
        "test_evidence": "tests/geo_ring_cloud_test_claas3.py::PackageBoundaryTests,RegistryTests",
        "notes": "Stable source IDs must remain backward compatible.",
    },
    {
        "project_id": PROJECT_ID,
        "canonical_module": "geo_ring_cloud.lineage",
        "canonical_path": "third_report/code/geo_ring_cloud_stage1/geo_ring_cloud/lineage.py",
        "component_role": "lineage",
        "legacy_module": "geo_ring_cloud_lineage",
        "legacy_path": "third_report/code/geo_ring_cloud_stage1/geo_ring_cloud_lineage.py",
        "migration_status": "canonical_with_compatibility_shim",
        "public_api": "utc_now, code_commit, write_manifest",
        "test_evidence": "tests/geo_ring_cloud_test_claas3.py::PackageBoundaryTests",
        "notes": "Manifest schema remains compatible with existing stage outputs.",
    },
    {
        "project_id": PROJECT_ID,
        "canonical_module": "geo_ring_cloud.run_discovery",
        "canonical_path": "third_report/code/geo_ring_cloud_stage1/geo_ring_cloud/run_discovery.py",
        "component_role": "run_discovery",
        "legacy_module": "geo_ring_cloud_run_discovery",
        "legacy_path": "third_report/code/geo_ring_cloud_stage1/geo_ring_cloud_run_discovery.py",
        "migration_status": "canonical_with_compatibility_shim",
        "public_api": "matrix_manifests, resolve_run_dir, discover_run_dirs, run_time_tag",
        "test_evidence": "tests/geo_ring_cloud_test_claas3.py::PackageBoundaryTests,RunDiscoveryTests",
        "notes": "Supports canonical matrix manifests and legacy time-tag directories.",
    },
    {
        "project_id": PROJECT_ID,
        "canonical_module": "geo_ring_cloud.pipeline_layout",
        "canonical_path": "third_report/code/geo_ring_cloud_stage1/geo_ring_cloud/pipeline_layout.py",
        "component_role": "pipeline_layout",
        "legacy_module": "stage1_common",
        "legacy_path": "third_report/code/geo_ring_cloud_stage1/stage1_common.py",
        "migration_status": "canonical_extracted",
        "public_api": "Stage 1 input/output paths, PIPELINE_DIRECTORIES, ensure_pipeline_directories",
        "test_evidence": "tests/geo_ring_cloud_test_claas3.py::PackageBoundaryTests,PipelineLayoutTests",
        "notes": "Filesystem layout is isolated from product reading and scientific semantics.",
    },
    {
        "project_id": PROJECT_ID,
        "canonical_module": "geo_ring_cloud.cloud_semantics",
        "canonical_path": "third_report/code/geo_ring_cloud_stage1/geo_ring_cloud/cloud_semantics.py",
        "component_role": "cloud_semantics",
        "legacy_module": "stage1_common",
        "legacy_path": "third_report/code/geo_ring_cloud_stage1/stage1_common.py",
        "migration_status": "canonical_extracted",
        "public_api": "cloud code tables, display/fusion masks, validity masks and quality normalization",
        "test_evidence": "tests/geo_ring_cloud_test_claas3.py::PackageBoundaryTests,CloudSemanticsTests",
        "notes": "Canonical scientific semantics used directly by Stage 05 and Stage 06.",
    },
    {
        "project_id": PROJECT_ID,
        "canonical_module": "geo_ring_cloud.fusion_support",
        "canonical_path": "third_report/code/geo_ring_cloud_stage1/geo_ring_cloud/fusion_support.py",
        "component_role": "fusion_support",
        "legacy_module": "06_fuse_best_source",
        "legacy_path": "third_report/code/geo_ring_cloud_stage1/06_fuse_best_source.py",
        "migration_status": "canonical_extracted",
        "public_api": "reprojected catalog, target grid, GEO geometry weights, cloud mapping, fusion candidate construction",
        "test_evidence": "tests/geo_ring_cloud_test_claas3.py::PackageBoundaryTests,FusionAndOverlapSupportTests",
        "notes": "Stage 06/06c/07 consumers use a static package API instead of dynamically loading the Stage 06 script.",
    },
    {
        "project_id": PROJECT_ID,
        "canonical_module": "geo_ring_cloud.reprojection",
        "canonical_path": "third_report/code/geo_ring_cloud_stage1/geo_ring_cloud/reprojection.py",
        "component_role": "reprojection",
        "legacy_module": "05_reproject_cloud_to_grid",
        "legacy_path": "third_report/code/geo_ring_cloud_stage1/05_reproject_cloud_to_grid.py",
        "migration_status": "canonical_extracted",
        "public_api": "target grid, native geolocation, coordinate normalization, KD-tree nearest-neighbor reprojection",
        "test_evidence": "tests/geo_ring_cloud_test_claas3.py::PackageBoundaryTests,ReprojectionAndGeometrySupportTests",
        "notes": "Stage 05 and Stage 06e share one static geolocation/reprojection implementation.",
    },
    {
        "project_id": PROJECT_ID,
        "canonical_module": "geo_ring_cloud.geometry",
        "canonical_path": "third_report/code/geo_ring_cloud_stage1/geo_ring_cloud/geometry.py",
        "component_role": "geometry",
        "legacy_module": "06c_geometry_parameter_audit",
        "legacy_path": "third_report/code/geo_ring_cloud_stage1/06c_geometry_parameter_audit.py",
        "migration_status": "canonical_extracted",
        "public_api": "GEO geometry parameters, audited subpoints, spherical and ellipsoidal ECEF VZA",
        "test_evidence": "tests/geo_ring_cloud_test_claas3.py::PackageBoundaryTests,ReprojectionAndGeometrySupportTests",
        "notes": "Stage 06c/06e consumers no longer load geometry implementation from a stage script path.",
    },
    {
        "project_id": PROJECT_ID,
        "canonical_module": "geo_ring_cloud.overlap",
        "canonical_path": "third_report/code/geo_ring_cloud_stage1/geo_ring_cloud/overlap.py",
        "component_role": "overlap_metrics",
        "legacy_module": "07_overlap_consistency_validation",
        "legacy_path": "third_report/code/geo_ring_cloud_stage1/07_overlap_consistency_validation.py",
        "migration_status": "canonical_extracted",
        "public_api": "source-variable mapping, binary confusion, source-boundary metrics, overlap quicklook",
        "test_evidence": "tests/geo_ring_cloud_test_claas3.py::PackageBoundaryTests,FusionAndOverlapSupportTests",
        "notes": "Stage 07 and 07p share one implementation without dynamically importing each other.",
    },
    {
        "project_id": PROJECT_ID,
        "canonical_module": "geo_ring_cloud.data_asset_audit",
        "canonical_path": "third_report/code/geo_ring_cloud_stage1/geo_ring_cloud/data_asset_audit.py",
        "component_role": "data_asset_audit",
        "legacy_module": "06f_reexport_with_obitype_patch",
        "legacy_path": "third_report/code/geo_ring_cloud_stage1/06f_reexport_with_obitype_patch.py",
        "migration_status": "canonical_extracted",
        "public_api": "apply_fy4b_obitype_patch",
        "test_evidence": "tests/geo_ring_cloud_test_claas3.py::DataAssetAuditTests",
        "notes": "Semantic database correction is testable independently; the legacy re-export script is process orchestration only.",
    },
    {
        "project_id": PROJECT_ID,
        "canonical_module": "geo_ring_cloud.adapters.cloud_products",
        "canonical_path": "third_report/code/geo_ring_cloud_stage1/geo_ring_cloud/adapters/cloud_products.py",
        "component_role": "product_adapter",
        "legacy_module": "stage1_common",
        "legacy_path": "third_report/code/geo_ring_cloud_stage1/stage1_common.py",
        "migration_status": "canonical_extracted",
        "public_api": "generic NetCDF/HDF/GRIB product readers, variable mapping, unit normalization, Himawari R21 geometry",
        "test_evidence": "tests/geo_ring_cloud_test_claas3.py::PackageBoundaryTests,CloudProductAdapterTests",
        "notes": "Format adaptation is isolated from rendering, serialization, layout and stage orchestration.",
    },
    {
        "project_id": PROJECT_ID,
        "canonical_module": "geo_ring_cloud.artifact_io",
        "canonical_path": "third_report/code/geo_ring_cloud_stage1/geo_ring_cloud/artifact_io.py",
        "component_role": "artifact_io",
        "legacy_module": "stage1_common",
        "legacy_path": "third_report/code/geo_ring_cloud_stage1/stage1_common.py",
        "migration_status": "canonical_extracted",
        "public_api": "safe_name, write_json_npz",
        "test_evidence": "tests/geo_ring_cloud_test_claas3.py::PackageBoundaryTests,ArtifactIoTests",
        "notes": "Preserves the established Stage 1 NPZ arrays plus JSON metadata schema.",
    },
    {
        "project_id": PROJECT_ID,
        "canonical_module": "geo_ring_cloud.quicklooks",
        "canonical_path": "third_report/code/geo_ring_cloud_stage1/geo_ring_cloud/quicklooks.py",
        "component_role": "quicklook_renderer",
        "legacy_module": "stage1_common",
        "legacy_path": "third_report/code/geo_ring_cloud_stage1/stage1_common.py",
        "migration_status": "canonical_extracted",
        "public_api": "make_quicklook",
        "test_evidence": "tests/geo_ring_cloud_test_claas3.py::PackageBoundaryTests,QuicklookTests",
        "notes": "Bounded-memory rendering with separate categorical and continuous display behavior.",
    },
    {
        "project_id": PROJECT_ID,
        "canonical_module": "geo_ring_cloud.pipeline_support",
        "canonical_path": "third_report/code/geo_ring_cloud_stage1/geo_ring_cloud/pipeline_support.py",
        "component_role": "compatibility_facade",
        "legacy_module": "stage1_common",
        "legacy_path": "third_report/code/geo_ring_cloud_stage1/stage1_common.py",
        "migration_status": "canonical_compatibility_facade",
        "public_api": "re-export facade for historical Stage 1 shared APIs",
        "test_evidence": "tests/geo_ring_cloud_test_claas3.py::PackageBoundaryTests",
        "notes": "Pure imports and aliases only; governance rejects implementation logic and active stage callers use focused modules.",
    },
    {
        "project_id": PROJECT_ID,
        "canonical_module": "geo_ring_cloud.diagnostics.summary",
        "canonical_path": "third_report/code/geo_ring_cloud_stage1/geo_ring_cloud/diagnostics/summary.py",
        "component_role": "diagnostics_library",
        "legacy_module": "stage1_common",
        "legacy_path": "third_report/code/geo_ring_cloud_stage1/stage1_common.py",
        "migration_status": "canonical_extracted",
        "public_api": "finite_stats",
        "test_evidence": "tests/geo_ring_cloud_test_claas3.py::SummaryDiagnosticsTests",
        "notes": "Deterministic compact array statistics shared by validation and fusion stages.",
    },
    {
        "project_id": PROJECT_ID,
        "canonical_module": "geo_ring_cloud.adapters.claas3",
        "canonical_path": "third_report/code/geo_ring_cloud_stage1/geo_ring_cloud/adapters/claas3.py",
        "component_role": "product_adapter",
        "legacy_module": "geo_ring_cloud_claas3_adapter",
        "legacy_path": "third_report/code/geo_ring_cloud_stage1/geo_ring_cloud_claas3_adapter.py",
        "migration_status": "canonical_with_compatibility_shim",
        "public_api": "CLAAS-3 file discovery, time selection, structure signature, normalized product read",
        "test_evidence": "tests/geo_ring_cloud_test_claas3.py::PackageBoundaryTests,Claas3AdapterTests",
        "notes": "CMA/CTX/CPP reader and normalization contract; legacy import remains a pure shim.",
    },
    {
        "project_id": PROJECT_ID,
        "canonical_module": "geo_ring_cloud.adapters.epic",
        "canonical_path": "third_report/code/geo_ring_cloud_stage1/geo_ring_cloud/adapters/epic.py",
        "component_role": "product_adapter",
        "legacy_module": "",
        "legacy_path": "third_report/code/geo_ring_cloud_stage1/stage_10_cth_validation/stage_10_run_cth_validation.py",
        "migration_status": "canonical_extracted",
        "public_api": "EPIC_CTH_CANDIDATES, read_epic_cth",
        "test_evidence": "tests/geo_ring_cloud_test_claas3.py::EpicAdapterTests",
        "notes": "Removes diagnostics-to-stage reverse dependency; Stage 10 retains a compatibility wrapper.",
    },
    {
        "project_id": PROJECT_ID,
        "canonical_module": "geo_ring_cloud.diagnostics.epic_pair",
        "canonical_path": "third_report/code/geo_ring_cloud_stage1/geo_ring_cloud/diagnostics/epic_pair.py",
        "component_role": "diagnostics_library",
        "legacy_module": "geo_ring_cloud_epic_pair_diagnostics",
        "legacy_path": "third_report/code/geo_ring_cloud_stage1/geo_ring_cloud_epic_pair_diagnostics.py",
        "migration_status": "canonical_with_compatibility_shim",
        "public_api": "EPIC/GEO paired sampling, strata, classification and height metrics, bootstrap",
        "test_evidence": "tests/geo_ring_cloud_test_claas3.py::PackageBoundaryTests,DiagnosticTests",
        "notes": "Reusable diagnostics depend on package adapters and source registry, not stage scripts.",
    },
)
INDEX_EXCLUDED_PARTS = {"__pycache__", ".pytest_cache", "_tmp"}
COMPONENT_ROLE_ASSIGNMENT = re.compile(
    r"^\s*COMPONENT_ROLE\s*=\s*['\"]([a-z][a-z0-9_]*)['\"]",
    re.MULTILINE,
)
TIME_RUN_STAGE_ROOT = re.compile(r"^stage_?\d{2}[a-z0-9_]*", re.IGNORECASE)
LEGACY_STAGE_FILENAME = re.compile(r"(?:^|[_-])(?:step|stage)\d{1,2}[a-z0-9]*(?:[_-]|$)", re.IGNORECASE)
CANONICAL_STAGE_FILENAME = re.compile(r"(?:^|[_-])stage_\d{2}(?:_[0-9]+|[a-z0-9]+|_[a-z0-9]+)?(?:[_-]|$)", re.IGNORECASE)
TIME_RUN_KEYWORDS = (
    "audit",
    "decision",
    "index",
    "inventory",
    "manifest",
    "metrics",
    "readme",
    "report",
    "status",
    "summary",
    "warning",
)
TIME_RUN_ARTIFACT_ROLES = {
    "experiment_runner",
    "presentation_builder",
    "profile_matrix_run",
    "stage_run_artifact",
    "time_run_pruning",
}
ARTIFACT_STAGE_HINTS = {
    "stage_00d_claas3": "stage_00d",
    "stage_06c_claas3": "stage_06c",
    "stage_07p_claas3": "stage_07p",
    "stage_09d_claas3": "stage_09d",
    "stage_10_claas3": "stage_10",
    "stage_10p2": "stage_10p2",
    "stage_10p": "stage_10p",
    "stage_10_cth_validation": "stage_10",
    "stage_10_": "stage_10",
    "stage_09f": "stage_09f",
    "stage_09e": "stage_09e",
    "core_time_index": "stage_01",
    "standardized_native_build": "stage_02",
    "standardized_native_validate": "stage_03",
    "semantic_validation": "stage_03_5",
    "fy4b_geo_alignment": "stage_04",
    "fy4b_dqf": "stage_04b",
    "reproject_cloud_to_grid": "stage_05",
    "fuse_best_source": "stage_06",
    "cloud_mask_fusion": "stage_06",
    "source_selection": "stage_06_5",
    "geometry_parameter": "stage_06c",
    "geometry_metadata": "stage_06c",
    "vza_method": "stage_06c",
    "himawari_full_disk": "stage_06d",
    "angle_source": "stage_06e",
    "unknown_aware": "stage_06f",
    "data_asset_audit": "stage_06f",
    "overlap_validation_07p": "stage_07p",
    "overlap_validation_report": "stage_07",
    "boundary_magnitude": "stage_07p_b",
    "single_time_acceptance": "stage_07v2",
    "epic_visual": "stage_08",
    "epic_l2_cloud": "stage_08b",
    "semantic_sensitivity": "stage_08c",
    "target_selection": "stage_08d",
    "multisample_summary": "stage_08e",
    "geometry_prefusion": "stage_08f",
    "overlap_count": "stage_08g",
    "time_offset": "stage_08h",
    "source_overlap": "stage_08i",
    "prefusion_source_pair": "stage_08j",
    "stage08_epic_comparison": "stage_08k",
    "stage09b": "stage_09b",
    "stage09c": "stage_09c",
    "stage09d": "stage_09d",
    "stage09": "stage_09",
}

DATA_PRODUCT_AUDIT_OVERRIDES = {
    "audit_geometry_and_variables.py": ("GEO cloud products", "geometry and variable inventory", "stage_00c", "generic product capability audit"),
    "check_download_completeness.py": ("GEO raw/cloud products", "download completeness", "stage_00", "data availability audit"),
    "claas3_202403_audit.py": ("CLAAS3", "structure and availability", "stage_00d", "Meteosat upstream product audit"),
    "compare_claas3_and_operational_meteosat.py": ("CLAAS3 / operational Meteosat", "cross-product comparison", "stage_00d", "Meteosat product family comparison"),
    "deep_claas3_structure_audit.py": ("CLAAS3", "deep product structure", "stage_00d", "Meteosat product structure audit"),
    "fy4b_cpd_product_inspection.py": ("FY4B CPD", "product structure, variables, flags, quicklook readiness", "", "generic FY4B CPD product inspection"),
    "latest_fy4b_goes_structure_probe.py": ("FY4B / GOES", "latest structure probe", "stage_00b,stage_00c", "cross-satellite sample probe"),
    "make_claas3_operational_visual_check_20240305.py": ("CLAAS3 / operational Meteosat", "visual check", "stage_00d", "representative visual audit"),
    "manual_variable_mapping_by_product.yaml": ("GEO cloud products", "manual variable mapping", "stage_00c,stage_02", "audit-derived mapping config"),
    "meteosat_catalogue_discovery.py": ("Meteosat", "catalogue discovery", "stage_00d", "upstream catalogue audit"),
    "meteosat_priority_cloud_download.py": ("Meteosat", "priority download", "stage_00f", "download planning utility"),
    "meteosat_product_series_audit.py": ("Meteosat", "product series audit", "stage_00d", "product series discovery"),
    "probe_claas3_downloaded_data.py": ("CLAAS3", "downloaded data probe", "stage_00d", "local downloaded product probe"),
    "probe_meteosat_grib_cfgrib.py": ("Meteosat GRIB", "cfgrib read probe", "stage_00d", "format read probe"),
    "read_one_sample_each_product.py": ("GEO cloud products", "one-sample read", "stage_00b", "sample-level product read audit"),
    "rewrite_meteosat_catalogue_report.py": ("Meteosat", "catalogue report rewrite", "stage_00d", "report postprocess"),
    "standardize_one_time_cloud_v0.py": ("GEO cloud products", "one-time standardization prototype", "stage_00e", "prototype standardization audit"),
}

STAGE_SCOPED_DATA_PRODUCT_AUDITS = [
    {
        "audit_id": "stage_00d_claas3_integration_readiness",
        "primary_path": str(ROOT / "third_report/code/geo_ring_cloud_stage1/stage_00d_claas3_integration_readiness.py"),
        "canonical_stage_id": "stage_00d",
        "related_stage_ids": "stage_00d,stage_01,stage_02,stage_03_5",
        "data_domain": "Meteosat 0deg",
        "product_family": "CM SAF CLAAS-3 V003 ICDR CMA/CTX/CPP",
        "audit_scope": "March 2024 cadence, duplicate/version resolution, structure, CF projection, QA flags, and integration contract",
        "status": "active",
        "output_root": str(ROOT / "data_check_report/claas3_integration_readiness"),
        "notes": "PASS_WITH_WARNINGS on local data; warnings are deterministic duplicate-order records",
    },
    {
        "audit_id": "stage_10p_epic_composite_psf_inventory",
        "primary_path": str(ROOT / "third_report/code/geo_ring_cloud_stage1/stage_10p_composite_inventory.py"),
        "canonical_stage_id": "stage_10p",
        "related_stage_ids": "stage_10,stage_10p",
        "data_domain": "DSCOVR EPIC",
        "product_family": "EPIC L2 COMPOSITE_02 / PSF-aware benchmark evidence",
        "audit_scope": "Composite file, variable, global-attribute, keyword, and cloud-property candidate inventory",
        "status": "active",
        "output_root": str(ROOT / "geo_ring_cloud_stage1_time_runs/stage_10p_psf_inventory_202401"),
        "notes": "Primary role is data_product_audit; related to Stage 10 validation but not a production pipeline transform",
    },
]


def canonical_stage_id(label: str) -> str:
    """Return canonical stage id for a legacy stage label, or empty for non-stage roles."""
    if not label or label in COMPONENT_ROLES:
        return ""
    value = label.strip().lower().replace("stage", "").replace("step", "")
    value = value.replace(".", "_").replace("-", "_")
    return f"stage_{value}" if value.startswith("0") else f"stage_{value.zfill(2)}"


def display_stage(canonical_id: str) -> str:
    if not canonical_id:
        return ""
    value = canonical_id.replace("stage_", "")
    return "Stage " + value.replace("_", ".")


def stage_sort_key(canonical_id: str) -> str:
    return canonical_id.replace("stage_", "")


def legacy_labels_for(label: str, script_name: str = "") -> str:
    if not label:
        return ""
    labels = {label}
    compact = label.replace(".", "").replace("-", "")
    labels.add(f"Stage{label}")
    labels.add(f"stage{label}")
    labels.add(f"Stage{compact}")
    labels.add(f"stage{compact}")
    labels.add(f"stage_{label.replace('.', '_').replace('-', '_')}")
    if script_name:
        labels.add(script_name)
    return ",".join(sorted(labels))


def classify_artifact(path: Path) -> str:
    if path.is_dir():
        return "directory_summary"
    ext = path.suffix.lower()
    if ext == ".py":
        return "script"
    if ext == ".ps1":
        return "script_helper"
    if ext == ".md":
        return "report"
    if ext == ".csv":
        return "table"
    if ext in (".json", ".yaml", ".yml", ".toml"):
        return "config_or_manifest"
    if ext == ".xlsx":
        return "workbook"
    return "other"


def infer_canonical_from_path(path: Path) -> tuple[str, str, str]:
    """Infer project, canonical stage, and legacy stage label from a path."""
    text = str(path).replace("\\", "/").lower()
    name = path.name.lower()
    if "/third_report/code/geo_data_audit/" in text or text.endswith("/third_report/code/geo_data_audit"):
        return PROJECT_ID, "", "data_product_audit"
    if "_non_geo_archive" in text and "epic_ceres" in text:
        if "stage9_5" in text or "stage_9_5" in text or "run_stage_9_5" in text:
            return "epic_ceres", "stage_09_5", "epic_ceres_stage9_5"
        if "stage9" in text or "run_stage_6_9" in text:
            return "epic_ceres", "stage_09", "epic_ceres_stage9"
        return "epic_ceres", "", ""
    if "claas3_stage0910" in text or "geo_ring_cloud_profile_pair" in text or "/experiments/" in text:
        return PROJECT_ID, "", "experiment_runner"
    if "geo_ring_cloud_stage1_time_runs/stage09d" in text or "stage09d" in name:
        return PROJECT_ID, "stage_09d", "stage09d"
    if "geo_ring_cloud_stage1_time_runs/stage09c" in text or "stage09c" in name:
        return PROJECT_ID, "stage_09c", "stage09c"
    if "geo_ring_cloud_stage1_time_runs/stage09b" in text or "stage09b" in name:
        return PROJECT_ID, "stage_09b", "stage09b"
    if "stage09_epic_georing_cloud_mask_diagnostics" in text or "stage09_" in name:
        return PROJECT_ID, "stage_09", "stage09"
    for token, canonical in ARTIFACT_STAGE_HINTS.items():
        if token in text:
            return PROJECT_ID, canonical, token
    if "/pipeline_stages/" in text:
        for stage, _, _, _, _, _ in PIPELINE_STAGES:
            compact = stage.replace(".", "_").replace("-", "_")
            if f"/{compact}_" in text or f"/stage_{compact}" in text:
                return PROJECT_ID, canonical_stage_id(stage), stage
    for stage, _, _, _, _, _ in PIPELINE_STAGES:
        canonical = canonical_stage_id(stage)
        if not canonical:
            continue
        token = canonical.replace("stage_", "")
        candidates = {
            f"/{token}_",
            f"/{token.replace('_', '.')}_",
            f"stage{token.replace('_', '')}",
            f"_{token}_",
            f"_{token.replace('_', '.')}_",
        }
        if any(c in text for c in candidates):
            return PROJECT_ID, canonical, stage
    return PROJECT_ID, "", ""


def component_role_for_path(path: Path, canonical: str, legacy: str) -> str:
    text = str(path).replace("\\", "/").lower()
    name = path.name.lower()
    if "geo_ring_cloud_time_run_pruning_" in text:
        return "time_run_pruning"
    if "claas3_stage0910" in text or "geo_ring_cloud_profile_pair" in text or "/experiments/" in text:
        return "experiment_runner"
    if "/presentations/" in text or text.endswith("/presentations"):
        return "presentation_builder"
    if "claas3_epic_" in text and "time_runs" in text:
        return "profile_matrix_run"
    if "/third_report/code/geo_data_audit/" in text or text.endswith("/third_report/code/geo_data_audit"):
        return "data_product_audit"
    if name == "stage_10p_composite_inventory.py" or "stage_10p_psf_inventory" in text:
        return "data_product_audit"
    return "" if canonical else COMPONENT_ROLES.get(legacy, "")


def declared_component_role(path: Path) -> str:
    try:
        text = path.read_text(encoding="utf-8")
    except (OSError, UnicodeDecodeError):
        return ""
    match = COMPONENT_ROLE_ASSIGNMENT.search(text)
    return match.group(1) if match else ""


def component_role_for_script(path: Path, rel_name: str, canonical: str) -> str:
    normalized = rel_name.replace("\\", "/").lower()
    if normalized == "stage_10p_composite_inventory.py":
        return "data_product_audit"
    declared = declared_component_role(path)
    return declared or ("" if canonical else "support")


def is_index_excluded(path: Path) -> bool:
    return any(part.lower() in INDEX_EXCLUDED_PARTS for part in path.parts)


def time_run_root_role(path: Path) -> str:
    name = path.name.lower()
    if name.startswith("geo_ring_cloud_time_run_pruning_"):
        return "time_run_pruning"
    if name == "experiments" or name.startswith("geo_ring_cloud_profile_pair_"):
        return "experiment_runner"
    if name == "presentations":
        return "presentation_builder"
    if TIME_RUN_STAGE_ROOT.match(name):
        return "stage_run_artifact"
    return ""


def should_index_time_run_artifact(path: Path, root: Path) -> bool:
    if is_index_excluded(path) or path.suffix.lower() not in KEY_ARTIFACT_EXTS:
        return False
    rel = path.relative_to(root)
    lowered_parts = {part.lower() for part in rel.parts}
    if lowered_parts & {"checkpoints", "diagnostic_cache", "pruning", "samples"}:
        return False
    return len(rel.parts) <= 3 or any(token in path.name.lower() for token in TIME_RUN_KEYWORDS)

# ---------------------------------------------------------------------------
# 7. 主流程：扫描 → 建 sqlite → 导出 xlsx
# ---------------------------------------------------------------------------

def create_schema(conn):
    cur = conn.cursor()
    cur.executescript("""
    DROP TABLE IF EXISTS directories;
    DROP TABLE IF EXISTS files;
    DROP TABLE IF EXISTS scripts;
    DROP TABLE IF EXISTS module_registry;
    DROP TABLE IF EXISTS external_data_refs;
    DROP TABLE IF EXISTS pipeline_stages;
    DROP TABLE IF EXISTS time_runs;
    DROP TABLE IF EXISTS external_disks;
    DROP TABLE IF EXISTS stage_registry;
    DROP TABLE IF EXISTS stage_aliases;
    DROP TABLE IF EXISTS artifact_index;
    DROP TABLE IF EXISTS data_product_audits;
    DROP TABLE IF EXISTS naming_violations;
    DROP TABLE IF EXISTS meta;

    CREATE TABLE directories (
        id INTEGER PRIMARY KEY,
        path TEXT, name TEXT, parent TEXT,
        relevance TEXT, role TEXT,
        file_count INTEGER, size_bytes INTEGER, size_text TEXT,
        ext_summary TEXT, referenced_by_code TEXT, note TEXT,
        last_seen TEXT, exists_now TEXT, move_candidate TEXT, path_risk TEXT
    );
    CREATE TABLE files (
        id INTEGER PRIMARY KEY,
        path TEXT, name TEXT, dir_path TEXT,
        ext TEXT, category TEXT, stage TEXT, relevance TEXT,
        size_bytes INTEGER, note TEXT
    );
    CREATE TABLE scripts (
        id INTEGER PRIMARY KEY,
        path TEXT, filename TEXT, stage TEXT, project_id TEXT,
        canonical_stage_id TEXT, component_role TEXT, legacy_stage TEXT, responsibility TEXT,
        refs_external_paths TEXT
    );
    CREATE TABLE module_registry (
        id INTEGER PRIMARY KEY,
        project_id TEXT,
        canonical_module TEXT,
        canonical_path TEXT,
        component_role TEXT,
        legacy_module TEXT,
        legacy_path TEXT,
        migration_status TEXT,
        public_api TEXT,
        test_evidence TEXT,
        notes TEXT,
        UNIQUE(project_id, canonical_module)
    );
    CREATE TABLE external_data_refs (
        id INTEGER PRIMARY KEY,
        external_path TEXT, referenced_by_script TEXT, line_no TEXT,
        purpose TEXT, location TEXT
    );
    CREATE TABLE pipeline_stages (
        id INTEGER PRIMARY KEY,
        stage TEXT, name TEXT, input TEXT, output TEXT,
        gate_status TEXT, evidence_dir TEXT
    );
    CREATE TABLE time_runs (
        id INTEGER PRIMARY KEY,
        run_id TEXT, kind TEXT, target_utc TEXT, note TEXT,
        parent_run_id TEXT, source_profile TEXT, status TEXT, manifest_path TEXT
    );
    CREATE TABLE external_disks (
        id INTEGER PRIMARY KEY,
        path TEXT, location TEXT, description TEXT, referenced TEXT
    );
    CREATE TABLE stage_registry (
        id INTEGER PRIMARY KEY,
        project_id TEXT,
        canonical_stage_id TEXT,
        display_label TEXT,
        legacy_labels TEXT,
        component_role TEXT,
        stage_order TEXT,
        name TEXT,
        meaning TEXT,
        input TEXT,
        output TEXT,
        status TEXT,
        evidence_paths TEXT,
        do_not_merge_with TEXT,
        notes TEXT,
        UNIQUE(project_id, canonical_stage_id)
    );
    CREATE TABLE stage_aliases (
        id INTEGER PRIMARY KEY,
        project_id TEXT,
        alias TEXT,
        canonical_stage_id TEXT,
        confidence REAL,
        reason TEXT
    );
    CREATE TABLE artifact_index (
        id INTEGER PRIMARY KEY,
        project_id TEXT,
        canonical_stage_id TEXT,
        component_role TEXT,
        artifact_type TEXT,
        path TEXT,
        legacy_stage_label TEXT,
        role TEXT,
        size_bytes INTEGER,
        file_count INTEGER,
        ext_summary TEXT,
        last_modified TEXT,
        summary TEXT,
        trace_source TEXT
    );
    CREATE TABLE data_product_audits (
        id INTEGER PRIMARY KEY,
        project_id TEXT,
        audit_id TEXT,
        primary_path TEXT,
        component_role TEXT,
        canonical_stage_id TEXT,
        related_stage_ids TEXT,
        data_domain TEXT,
        product_family TEXT,
        audit_scope TEXT,
        status TEXT,
        output_root TEXT,
        notes TEXT
    );
    CREATE TABLE naming_violations (
        id INTEGER PRIMARY KEY,
        project_id TEXT,
        path TEXT,
        legacy_label TEXT,
        issue_type TEXT,
        severity TEXT,
        suggested_canonical_stage_id TEXT,
        suggested_new_path TEXT,
        reason TEXT
    );
    CREATE TABLE meta (
        key TEXT PRIMARY KEY, value TEXT
    );
    CREATE INDEX idx_dirs_rel ON directories(relevance);
    CREATE INDEX idx_files_rel ON files(relevance);
    CREATE INDEX idx_files_cat ON files(category);
    CREATE INDEX idx_module_registry_role ON module_registry(project_id, component_role);
    CREATE INDEX idx_extref_loc ON external_data_refs(location);
    CREATE INDEX idx_stage_registry_project ON stage_registry(project_id);
    CREATE INDEX idx_artifact_stage ON artifact_index(project_id, canonical_stage_id);
    CREATE INDEX idx_data_product_audit_role ON data_product_audits(project_id, component_role);
    CREATE INDEX idx_alias_lookup ON stage_aliases(project_id, alias);
    """)
    conn.commit()


def categorize(name: str, ext: str) -> str:
    e = ext.lower()
    if e in (".py",):
        return "code"
    if e in (".ipynb",):
        return "notebook"
    if e in (".yaml", ".yml", ".json", ".toml", ".ini", ".cfg"):
        return "config"
    if e in (".md", ".rst", ".txt"):
        return "report"
    if e in (".csv",):
        return "table"
    if e in (".png", ".jpg", ".jpeg", ".svg", ".pdf"):
        return "quicklook"
    if e in (".npz", ".npy"):
        return "data_npz"
    if e in (".nc", ".nc4", ".hdf", ".h5", ".nat", ".dat", ".bz2", ".zip", ".rar", ".sqlite", ".db"):
        return "data"
    return "other"


def scan_files_in_dir(dirpath: Path, relevance: str, stage: str, conn):
    """全量入库一个目录下的所有文件（用于小文件目录：code/config/report/manifest）。"""
    cur = conn.cursor()
    if not dirpath.exists():
        return
    for dirpath_, dirnames, filenames in os.walk(dirpath):
        dirnames[:] = [d for d in dirnames if d not in
                       ("__pycache__", "node_modules", ".git", ".claude", "dist")]
        for f in filenames:
            if f.startswith("~$"):
                continue
            fp = Path(dirpath_) / f
            try:
                sz = fp.stat().st_size
            except OSError:
                sz = None
            ext = fp.suffix.lower()
            cat = categorize(f, ext)
            cur.execute(
                "INSERT INTO files(path,name,dir_path,ext,category,stage,relevance,size_bytes,note) "
                "VALUES(?,?,?,?,?,?,?,?,?)",
                (str(fp), f, str(dirpath_), ext, cat, stage, relevance, sz, ""),
            )
    conn.commit()


def summary_in_dir(dirpath: Path):
    """只统计不逐文件入库（用于海量数据目录：返回文件数/体积/扩展名汇总）。"""
    n, b, exts = dir_stats(dirpath)
    ext_summary = ", ".join(f"{k}:{v}" for k, v in sorted(exts.items(), key=lambda x: -x[1])[:6])
    return n, b, ext_summary


def insert_stage_registry(conn: sqlite3.Connection) -> None:
    cur = conn.cursor()
    rows: dict[tuple[str, str], dict[str, str]] = {}

    for stage, name, inp, outp, gate, evidence in PIPELINE_STAGES:
        canonical = canonical_stage_id(stage)
        rows[(PROJECT_ID, canonical)] = {
            "project_id": PROJECT_ID,
            "canonical_stage_id": canonical,
            "display_label": display_stage(canonical),
            "legacy_labels": legacy_labels_for(stage),
            "component_role": "",
            "stage_order": stage_sort_key(canonical),
            "name": name,
            "meaning": name,
            "input": inp,
            "output": outp,
            "status": gate,
            "evidence_paths": evidence,
            "do_not_merge_with": "epic_ceres.stage_09,epic_ceres.stage_09_5" if canonical == "stage_09" else "",
            "notes": "Canonical GEO-ring Cloud pipeline stage",
        }

    for fname, stage, responsibility in SCRIPTS:
        canonical = canonical_stage_id(stage)
        if canonical:
            key = (PROJECT_ID, canonical)
            existing = rows.get(key)
            if existing:
                existing["legacy_labels"] = ",".join(sorted(set((existing["legacy_labels"] + "," + legacy_labels_for(stage, fname)).split(","))))
                if fname not in existing["evidence_paths"]:
                    existing["evidence_paths"] = (existing["evidence_paths"] + "," + fname).strip(",")
            else:
                rows[key] = {
                    "project_id": PROJECT_ID,
                    "canonical_stage_id": canonical,
                    "display_label": display_stage(canonical),
                    "legacy_labels": legacy_labels_for(stage, fname),
                    "component_role": "",
                    "stage_order": stage_sort_key(canonical),
                    "name": responsibility.split("：", 1)[0],
                    "meaning": responsibility,
                    "input": "",
                    "output": "",
                    "status": "",
                    "evidence_paths": fname,
                    "do_not_merge_with": "epic_ceres.stage_09,epic_ceres.stage_09_5" if canonical == "stage_09" else "",
                    "notes": "Script-derived canonical stage not present in original pipeline table",
                }

    # Current post-Stage09 diagnostic code discovered from the file system must be first-class in the registry.
    for canonical, name, evidence in [
        ("stage_09b", "Stage 09b full overnight diagnostics", "geo_ring_cloud_stage1_time_runs/stage09b_full_202403_overnight_diagnostics"),
        ("stage_09c", "Stage 09c scaled March batch", "geo_ring_cloud_stage1_time_runs/stage09c_scaled_202403_batch"),
        ("stage_09d", "Stage 09d full-pixel diagnostics and interpretation", "geo_ring_cloud_stage1_time_runs/stage09d_full_pixel_diagnostics_202403"),
        ("stage_09e", "Stage 09e PSF-like EPIC-view spatial representativeness and SEL-QC diagnostics", "geo_ring_cloud_stage1_time_runs/stage_09e_psf_aware_epic_view_202403,geo_ring_cloud_stage1_time_runs/stage_09e_sel_qc_common_valid_202403"),
        ("stage_09f", "Stage 09F spatial story maps for GEO-ring vs EPIC cloud-mask diagnostics", "geo_ring_cloud_stage1_time_runs/stage_09f_spatial_story_maps_202403,stage_09f_spatial_story_maps/stage_09f_make_spatial_story_maps.py"),
        ("stage_10", "Stage 10 fused CTH validation and mechanism diagnostics", "geo_ring_cloud_stage1_time_runs/stage_10_cth_fused_product_validation_202403"),
        ("stage_10p", "Stage 10p related EPIC Composite PSF-aware data product audit", "geo_ring_cloud_stage1_time_runs/stage_10p_psf_inventory_202401"),
        ("stage_10p2", "Stage 10p2 approximate EPIC FOV aggregation diagnostics", "geo_ring_cloud_stage1_time_runs/stage_10p2_approx_epic_fov_aggregation_202403"),
    ]:
        key = (PROJECT_ID, canonical)
        collision_guard = "epic_ceres.stage_09,epic_ceres.stage_09_5" if canonical.startswith("stage_09") else ""
        rows.setdefault(key, {
            "project_id": PROJECT_ID,
            "canonical_stage_id": canonical,
            "display_label": display_stage(canonical),
            "legacy_labels": canonical.replace("stage_", "stage") + "," + canonical.replace("stage_", "Stage"),
            "component_role": "",
            "stage_order": stage_sort_key(canonical),
            "name": name,
            "meaning": name,
            "input": "Stage09 / Stage10 / time_runs outputs",
            "output": evidence,
            "status": "",
            "evidence_paths": evidence,
            "do_not_merge_with": collision_guard,
            "notes": "Current filesystem-derived diagnostic extension",
        })
        if canonical == "stage_10p":
            rows[key]["component_role"] = "data_product_audit"
            rows[key]["notes"] = "Stage-scoped data product audit; primary lookup is data_product_audits.md"
        if collision_guard:
            rows[key]["do_not_merge_with"] = collision_guard
        if evidence not in rows[key]["evidence_paths"]:
            rows[key]["evidence_paths"] = (rows[key]["evidence_paths"] + "," + evidence).strip(",")

    code_root = ROOT / "third_report/code/geo_ring_cloud_stage1"
    if code_root.exists():
        for fp in sorted(code_root.rglob("*.py")):
            if is_index_excluded(fp):
                continue
            rel_name = fp.relative_to(code_root).as_posix()
            stage = infer_stage_from_name(rel_name)
            canonical = canonical_stage_id(stage)
            if not canonical:
                continue
            key = (PROJECT_ID, canonical)
            if key not in rows:
                continue
            labels = set(filter(None, rows[key]["legacy_labels"].split(",")))
            labels.add(rel_name)
            rows[key]["legacy_labels"] = ",".join(sorted(labels))
            if rel_name not in rows[key]["evidence_paths"]:
                rows[key]["evidence_paths"] = (rows[key]["evidence_paths"] + "," + rel_name).strip(",")

    for project_id, canonical, name, legacy, notes in [
        ("epic_ceres", "stage_09", "EPIC-CERES Stage 09 statistics", "Stage9,stage9,Step9", "Collision guard only; not part of GEO-ring Cloud"),
        ("epic_ceres", "stage_09_5", "EPIC-CERES Stage 09.5 audit", "Stage9.5,stage9_5,run_stage_9_5_audit.py", "Collision guard only; not part of GEO-ring Cloud"),
    ]:
        rows[(project_id, canonical)] = {
            "project_id": project_id,
            "canonical_stage_id": canonical,
            "display_label": display_stage(canonical),
            "legacy_labels": legacy,
            "component_role": "",
            "stage_order": stage_sort_key(canonical),
            "name": name,
            "meaning": name,
            "input": "",
            "output": "",
            "status": "archived",
            "evidence_paths": str(ARCHIVE_DIR / "third_report"),
            "do_not_merge_with": "geo_ring_cloud.stage_09,geo_ring_cloud.stage_09b,geo_ring_cloud.stage_09c,geo_ring_cloud.stage_09d,geo_ring_cloud.stage_09e,geo_ring_cloud.stage_09f",
            "notes": notes,
        }

    for row in rows.values():
        cur.execute(
            "INSERT INTO stage_registry(project_id,canonical_stage_id,display_label,legacy_labels,component_role,stage_order,name,meaning,input,output,status,evidence_paths,do_not_merge_with,notes) "
            "VALUES(:project_id,:canonical_stage_id,:display_label,:legacy_labels,:component_role,:stage_order,:name,:meaning,:input,:output,:status,:evidence_paths,:do_not_merge_with,:notes)",
            row,
        )

    for row in rows.values():
        if not row["canonical_stage_id"]:
            continue
        aliases = [a for a in row["legacy_labels"].split(",") if a]
        aliases += [row["canonical_stage_id"], f"{row['project_id']}.{row['canonical_stage_id']}"]
        for alias in sorted(set(aliases)):
            cur.execute(
                "INSERT INTO stage_aliases(project_id,alias,canonical_stage_id,confidence,reason) VALUES(?,?,?,?,?)",
                (row["project_id"], alias, row["canonical_stage_id"], 1.0, "registry-defined alias"),
            )
    conn.commit()


def insert_artifact(conn: sqlite3.Connection, path: Path, role: str, trace_source: str, summary: str = "") -> None:
    cur = conn.cursor()
    project_id, canonical, legacy = infer_canonical_from_path(path)
    component_role = component_role_for_path(path, canonical, legacy)
    if path.is_file() and path.suffix.lower() == ".py":
        component_role = declared_component_role(path) or component_role
    try:
        stat = path.stat()
        size = stat.st_size if path.is_file() else None
        modified = datetime.fromtimestamp(stat.st_mtime, timezone.utc).isoformat().replace("+00:00", "Z")
    except OSError:
        size = None
        modified = ""
    file_count = None
    ext_summary = ""
    if path.is_dir():
        file_count, size, ext_summary = summary_in_dir(path)
    cur.execute(
        "INSERT INTO artifact_index(project_id,canonical_stage_id,component_role,artifact_type,path,legacy_stage_label,role,size_bytes,file_count,ext_summary,last_modified,summary,trace_source) "
        "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (
            project_id,
            canonical,
            component_role,
            classify_artifact(path),
            str(path),
            legacy,
            role,
            size,
            file_count,
            ext_summary,
            modified,
            summary or path.name,
            trace_source,
        ),
    )


def insert_artifacts(conn: sqlite3.Connection) -> None:
    roots = [
        (ROOT / "third_report/code/geo_ring_cloud_stage1", "source_code", "code root"),
        (ROOT / "third_report/code/geo_data_audit", "data_product_audit_source", "generic EO product audit code root"),
        (ROOT / "geo_ring_cloud_stage1/reports", "stage_report", "stage1 reports"),
        (ROOT / "geo_ring_cloud_stage1/config", "config", "stage1 config"),
        (ROOT / "geo_ring_cloud_stage1/time_index", "table", "stage1 time index"),
        (ROOT / "geo_ring_cloud_stage1_evidence_pack/latest", "evidence", "latest evidence pack"),
    ]
    time_runs = ROOT / "geo_ring_cloud_stage1_time_runs"
    if time_runs.exists():
        for d in sorted(time_runs.iterdir()):
            if not d.is_dir():
                continue
            run_role = time_run_root_role(d)
            if run_role:
                roots.append((d, run_role, f"time-runs {run_role}"))

    for root, role, trace in roots:
        if not root.exists():
            continue
        insert_artifact(conn, root, "directory_summary" if root.is_dir() else role, trace)
        if root.is_file():
            continue
        if role in TIME_RUN_ARTIFACT_ROLES:
            for child in sorted(root.iterdir()):
                if child.is_dir() and not is_index_excluded(child):
                    insert_artifact(conn, child, "directory_summary", f"{trace} section")
        for fp in sorted(root.rglob("*")):
            if not fp.is_file():
                continue
            if is_index_excluded(fp):
                continue
            if role in TIME_RUN_ARTIFACT_ROLES:
                include = should_index_time_run_artifact(fp, root)
            else:
                include = fp.suffix.lower() in KEY_ARTIFACT_EXTS
            if include:
                insert_artifact(conn, fp, role, trace)
    conn.commit()


def insert_data_product_audits(conn: sqlite3.Connection) -> None:
    cur = conn.cursor()
    audit_root = ROOT / "third_report/code/geo_data_audit"
    if audit_root.exists():
        for fp in sorted(audit_root.iterdir()):
            if not fp.is_file() or fp.suffix.lower() not in (".py", ".yaml", ".yml"):
                continue
            product_family, audit_scope, related, notes = DATA_PRODUCT_AUDIT_OVERRIDES.get(
                fp.name,
                ("Earth observation product", "generic product audit", "", "generic EO product audit utility"),
            )
            output_root = ""
            if fp.name == "fy4b_cpd_product_inspection.py":
                output_root = str(ROOT / "data_check_report/fy4b_cpd_product_inspection")
            elif fp.suffix.lower() == ".py":
                output_root = str(ROOT / "data_check_report")
            cur.execute(
                "INSERT INTO data_product_audits(project_id,audit_id,primary_path,component_role,canonical_stage_id,related_stage_ids,data_domain,product_family,audit_scope,status,output_root,notes) "
                "VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
                (
                    PROJECT_ID,
                    fp.stem,
                    str(fp),
                    "data_product_audit",
                    "",
                    related,
                    product_family.split()[0] if product_family else "Earth observation",
                    product_family,
                    audit_scope,
                    "active" if fp.exists() else "missing",
                    output_root,
                    notes,
                ),
            )

    for row in STAGE_SCOPED_DATA_PRODUCT_AUDITS:
        cur.execute(
            "INSERT INTO data_product_audits(project_id,audit_id,primary_path,component_role,canonical_stage_id,related_stage_ids,data_domain,product_family,audit_scope,status,output_root,notes) "
            "VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                PROJECT_ID,
                row["audit_id"],
                row["primary_path"],
                "data_product_audit",
                row["canonical_stage_id"],
                row["related_stage_ids"],
                row["data_domain"],
                row["product_family"],
                row["audit_scope"],
                row["status"],
                row["output_root"],
                row["notes"],
            ),
        )
    conn.commit()


def insert_naming_violations(conn: sqlite3.Connection) -> None:
    cur = conn.cursor()
    registered_legacy_interfaces = {
        row["legacy_path"].replace("\\", "/").lower()
        for row in MODULE_REGISTRY
        if row["legacy_module"] and row["legacy_path"]
    }
    for project_id, path, legacy, issue, severity, suggested, new_path, reason in [
        (PROJECT_ID, str(ROOT / "research_tracker"), "Step*/Stage* inferred labels", "untrusted_tracker_stage_inference", "warning", "", "", "research_tracker regex labels are not canonical taxonomy"),
        (PROJECT_ID, str(ROOT / "third_report/code/geo_ring_cloud_stage1"), "公共/运行器/下载/证据包/汇总", "component_role_misused_as_stage", "info", "", "", "These are component roles, not pipeline stages"),
        ("epic_ceres", str(ARCHIVE_DIR / "third_report/code/epic_ceres"), "Stage9/Stage9.5", "cross_project_stage_collision", "warning", "epic_ceres.stage_09 / epic_ceres.stage_09_5", "", "Do not merge archived EPIC-CERES stages with GEO-ring Cloud Stage09"),
    ]:
        cur.execute(
            "INSERT INTO naming_violations(project_id,path,legacy_label,issue_type,severity,suggested_canonical_stage_id,suggested_new_path,reason) VALUES(?,?,?,?,?,?,?,?)",
            (project_id, path, legacy, issue, severity, suggested, new_path, reason),
        )

    code_root = ROOT / "third_report/code/geo_ring_cloud_stage1"
    if code_root.exists():
        for fp in sorted(code_root.rglob("*")):
            if not fp.is_file() or is_index_excluded(fp):
                continue
            if fp.suffix.lower() not in {".py", ".ps1", ".json", ".yaml", ".yml"}:
                continue
            if not LEGACY_STAGE_FILENAME.search(fp.name) or CANONICAL_STAGE_FILENAME.search(fp.name):
                continue
            rel_name = fp.relative_to(code_root).as_posix()
            repo_relative = fp.relative_to(ROOT).as_posix().lower()
            if repo_relative in registered_legacy_interfaces:
                continue
            inferred = canonical_stage_id(infer_stage_from_name(rel_name))
            cur.execute(
                "INSERT INTO naming_violations(project_id,path,legacy_label,issue_type,severity,suggested_canonical_stage_id,suggested_new_path,reason) VALUES(?,?,?,?,?,?,?,?)",
                (
                    PROJECT_ID,
                    str(fp),
                    fp.name,
                    "legacy_source_filename",
                    "warning",
                    inferred,
                    "",
                    "Historical source name is not canonical; retain until imports, runners, evidence references, and rollback manifest are audited",
                ),
            )

    tracker_db = ROOT / "research_tracker" / "research_tracker.db"
    if tracker_db.exists():
        try:
            tc = sqlite3.connect(tracker_db)
            for stage, count in tc.execute("select stage,count(*) from files where stage like 'Step%' or stage like 'BuildStep%' or stage='Stage9' group by stage"):
                cur.execute(
                    "INSERT INTO naming_violations(project_id,path,legacy_label,issue_type,severity,suggested_canonical_stage_id,suggested_new_path,reason) VALUES(?,?,?,?,?,?,?,?)",
                    (PROJECT_ID, str(tracker_db), stage, "legacy_tracker_label", "warning", "", "", f"Observed {count} files with non-canonical tracker label; requires project-aware review"),
                )
            tc.close()
        except sqlite3.Error as exc:
            cur.execute(
                "INSERT INTO naming_violations(project_id,path,legacy_label,issue_type,severity,suggested_canonical_stage_id,suggested_new_path,reason) VALUES(?,?,?,?,?,?,?,?)",
                (PROJECT_ID, str(tracker_db), "", "tracker_db_unreadable", "warning", "", "", str(exc)),
            )
    conn.commit()


def build_sqlite():
    db_path = DB_PATH
    try:
        conn = sqlite3.connect(db_path)
        create_schema(conn)
    except sqlite3.Error:
        try:
            conn.close()
        except Exception:
            pass
        db_path = REFRESHED_DB_PATH
        if db_path.exists():
            db_path.unlink()
        conn = sqlite3.connect(db_path)
        create_schema(conn)
    cur = conn.cursor()

    # --- directories 表 ---
    for rel, relevance, role, size_note, refby, note in TOP_DIRS:
        full = ROOT / rel
        exists_now = full.exists()
        if full.is_file():
            # 单文件条目（如 pptx）
            try:
                n, b, ext_sum = 1, full.stat().st_size, full.suffix.lower()
            except OSError:
                n, b, ext_sum = 1, 0, ""
            parent = str(full.parent)
        else:
            n, b, ext_sum = summary_in_dir(full)
            parent = str(full.parent) if full.exists() else ""
        cur.execute(
            "INSERT INTO directories(path,name,parent,relevance,role,file_count,size_bytes,size_text,ext_summary,referenced_by_code,note,last_seen,exists_now,move_candidate,path_risk) "
            "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                str(full),
                full.name,
                parent,
                relevance,
                role,
                n,
                b,
                fmt_size(b) if b else size_note,
                ext_sum,
                refby,
                note,
                GENERATED_AT if exists_now else "",
                bool_text(exists_now),
                move_candidate(relevance, refby, exists_now),
                path_risk(relevance, refby, exists_now),
            ),
        )
    conn.commit()

    # --- files 表：全量入库小文件目录（代码/配置/报告/清单），数据目录只入库其下的脚本与文档 ---
    # 1) ring cloud 主代码（全量）
    scan_files_in_dir(ROOT / "third_report/code/geo_ring_cloud_stage1", "强相关", "", conn)
    # 2) geo_ring_cloud_stage1 产物根下的脚本/配置/报告/清单（全量，但跳过海量 npz/png 子目录的逐文件）
    #    此处只入库该目录直接子层及 reports/ scripts/ config/ time_index/ standardized_native 里的非数据小文件
    for sub in ["scripts", "config", "reports"]:
        d = ROOT / "geo_ring_cloud_stage1" / sub
        scan_files_in_dir(d, "强相关", "", conn)
    for sub in ["time_index"]:
        d = ROOT / "geo_ring_cloud_stage1" / sub
        scan_files_in_dir(d, "强相关", "01", conn)
    # 3) data_check_report 全量（都是审计报告/csv/yaml，小文件）
    scan_files_in_dir(ROOT / "data_check_report", "强相关", "00", conn)
    # 4) geo_geometry_check 全量（报告/csv + 少量样本）
    scan_files_in_dir(ROOT / "geo_geometry_check", "强相关", "06c", conn)
    # 5) evidence_pack 全量（md/json/csv/yaml）
    scan_files_in_dir(ROOT / "geo_ring_cloud_stage1_evidence_pack", "强相关", "", conn)
    # 6) third_report/code 上游模块（全量小文件）
    for sub in ["L1g", "FY4B", "GOES", "Himawari", "Meteosat", "geo_data_audit",
                "geo_cloud_download", "priority_download_goes_meteosat", "preview_baselines"]:
        scan_files_in_dir(ROOT / "third_report/code" / sub, "上游相关", "", conn)
    for f in ["standardized_l1_source_satpy.py", "run_standardized_l1_source_batch.py",
              "validate_standardized_l1_source_samples.py", "preview_runner.py",
              "plot_geo_satellite_coverage.py", "README.md",
              "project_progress_summary_20260505.md", "project_progress_summary_20260511.md"]:
        fp = ROOT / "third_report/code" / f
        if fp.is_file():
            cur.execute(
                "INSERT INTO files(path,name,dir_path,ext,category,stage,relevance,size_bytes,note) "
                "VALUES(?,?,?,?,?,?,?,?,?)",
                (str(fp), f, str(fp.parent), fp.suffix.lower(), categorize(f, fp.suffix.lower()),
                 "", "上游相关", fp.stat().st_size, ""))
    conn.commit()

    # --- scripts 表 ---
    code_root = ROOT / "third_report/code/geo_ring_cloud_stage1"
    # 计算每个脚本引用的外部路径
    refs_by_script = {}
    for ext_path, script, lineno, purpose, loc in EXT_REFS:
        refs_by_script.setdefault(script, []).append(ext_path)
    for fname, stage, resp in SCRIPTS:
        fp = code_root / fname
        refs = ",".join(sorted(set(refs_by_script.get(fname, []))))
        canonical = canonical_stage_id(stage)
        component_role = declared_component_role(fp) or ("" if canonical else COMPONENT_ROLES.get(stage, "support"))
        cur.execute(
            "INSERT INTO scripts(path,filename,stage,project_id,canonical_stage_id,component_role,legacy_stage,responsibility,refs_external_paths) "
            "VALUES(?,?,?,?,?,?,?,?,?)",
            (str(fp), fname, stage, PROJECT_ID, canonical, component_role, stage, resp, refs))
    known_script_names = {fname.replace("\\", "/") for fname, _, _ in SCRIPTS}
    if code_root.exists():
        for fp in sorted(code_root.rglob("*.py")):
            if is_index_excluded(fp):
                continue
            rel_name = fp.relative_to(code_root).as_posix()
            if rel_name in known_script_names:
                continue
            stage = infer_stage_from_name(rel_name)
            canonical = canonical_stage_id(stage)
            component_role = component_role_for_script(fp, rel_name, canonical)
            cur.execute(
                "INSERT INTO scripts(path,filename,stage,project_id,canonical_stage_id,component_role,legacy_stage,responsibility,refs_external_paths) "
                "VALUES(?,?,?,?,?,?,?,?,?)",
                (str(fp), rel_name, stage, PROJECT_ID, canonical, component_role, stage, summarize_script(rel_name), ""))
    conn.commit()

    # --- module_registry table ---
    for row in MODULE_REGISTRY:
        cur.execute(
            "INSERT INTO module_registry(project_id,canonical_module,canonical_path,component_role,legacy_module,legacy_path,migration_status,public_api,test_evidence,notes) "
            "VALUES(:project_id,:canonical_module,:canonical_path,:component_role,:legacy_module,:legacy_path,:migration_status,:public_api,:test_evidence,:notes)",
            row,
        )
    conn.commit()

    # --- external_data_refs 表 ---
    for ext_path, script, lineno, purpose, loc in EXT_REFS:
        cur.execute(
            "INSERT INTO external_data_refs(external_path,referenced_by_script,line_no,purpose,location) "
            "VALUES(?,?,?,?,?)",
            (ext_path, script, str(lineno), purpose, loc))
    # 外部盘
    for path, loc, desc, refby in EXTERNAL_DISKS:
        cur.execute(
            "INSERT INTO external_disks(path,location,description,referenced) VALUES(?,?,?,?)",
            (path, loc, desc, refby))
    conn.commit()

    # --- pipeline_stages 表 ---
    for stage, name, inp, outp, gate, edir in PIPELINE_STAGES:
        cur.execute(
            "INSERT INTO pipeline_stages(stage,name,input,output,gate_status,evidence_dir) "
            "VALUES(?,?,?,?,?,?)",
            (stage, name, inp, outp, gate, edir))
    conn.commit()

    # --- time_runs 表 ---
    for (rid,) in TIME_RUNS:
        # rid 如 20240308_0400 -> 2024-03-08T04:00Z
        try:
            d, h = rid.split("_")
            utc = f"{d[0:4]}-{d[4:6]}-{d[6:8]}T{h[0:2]}:{h[2:4]}Z"
        except Exception:
            utc = ""
        cur.execute("INSERT INTO time_runs(run_id,kind,target_utc,note) VALUES(?,?,?,?)",
                    (rid, "时间批次", utc, ""))
    for tid, desc in EPIC_TOPICS:
        cur.execute("INSERT INTO time_runs(run_id,kind,target_utc,note) VALUES(?,?,?,?)",
                    (tid, "EPIC专题", "", desc))
    matrix_root = ROOT / "geo_ring_cloud_stage1_time_runs"
    matrix_paths = sorted(matrix_root.glob("*/geo_ring_cloud_time_run_matrix_manifest.json")) if matrix_root.exists() else []
    for manifest_path in matrix_paths:
        try:
            matrix = json.loads(manifest_path.read_text(encoding="utf-8"))
        except Exception:
            continue
        common = matrix.get("common_inputs", {})
        for profile_run in matrix.get("profile_runs", []):
            cur.execute(
                "INSERT INTO time_runs(run_id,kind,target_utc,note,parent_run_id,source_profile,status,manifest_path) VALUES(?,?,?,?,?,?,?,?)",
                (
                    matrix.get("run_id", manifest_path.parent.name),
                    "PROFILE_MATRIX",
                    common.get("target_time", ""),
                    "CLAAS-3 matched profile run",
                    matrix.get("parent_run_id", matrix.get("run_id", "")),
                    profile_run.get("source_profile", ""),
                    profile_run.get("status", matrix.get("status", "")),
                    profile_run.get("manifest_path", ""),
                ),
            )
    conn.commit()

    # --- canonical taxonomy / memory index ---
    insert_stage_registry(conn)
    insert_artifacts(conn)
    insert_data_product_audits(conn)
    insert_naming_violations(conn)

    # --- meta 表 ---
    ts = GENERATED_AT
    meta = {
        "generated_at": ts,
        "project_root": str(ROOT),
        "schema_version": "2.0",
        "source_note": "由 build_index.py 基于人工相关性标签 + 当前文件系统扫描 + canonical taxonomy 生成；不移动项目数据",
        "task_name": "GEO-ring Cloud Stage1（多静止卫星统一云产品融合）",
        "prototype_time": "2024-03-05T00:00:00Z",
        "topology": "主代码 third_report/code/geo_ring_cloud_stage1 → 产物 geo_ring_cloud_stage1 → 多时次 time_runs → 证据包 evidence_pack",
        "workspace_dir": str(WORKSPACE_DIR),
        "archive_dir": str(ARCHIVE_DIR),
    }
    for k, v in meta.items():
        cur.execute("INSERT INTO meta(key,value) VALUES(?,?)", (k, v))
    conn.commit()
    conn.close()
    print(f"[OK] sqlite 写入: {db_path}")
    return db_path


def export_xlsx(db_path: Path = DB_PATH):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    conn = sqlite3.connect(db_path)
    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    wrap = Alignment(wrap_text=True, vertical="top")

    sheets = {
        "Canonical阶段注册表": "SELECT project_id,canonical_stage_id,component_role,display_label,name,status,legacy_labels,evidence_paths,do_not_merge_with,notes FROM stage_registry ORDER BY project_id,stage_order,canonical_stage_id",
        "阶段别名": "SELECT project_id,alias,canonical_stage_id,confidence,reason FROM stage_aliases ORDER BY project_id,alias",
        "产物索引": "SELECT project_id,canonical_stage_id,component_role,artifact_type,path,legacy_stage_label,role,size_bytes,file_count,ext_summary,last_modified,summary FROM artifact_index ORDER BY project_id,canonical_stage_id,artifact_type,path",
        "DataProductAudits": "SELECT project_id,audit_id,component_role,canonical_stage_id,related_stage_ids,data_domain,product_family,audit_scope,status,primary_path,output_root,notes FROM data_product_audits ORDER BY component_role,canonical_stage_id,audit_id",
        "命名问题": "SELECT project_id,path,legacy_label,issue_type,severity,suggested_canonical_stage_id,suggested_new_path,reason FROM naming_violations ORDER BY severity DESC,project_id,path",
        "目录相关性": "SELECT path,relevance,role,file_count,size_text,referenced_by_code,exists_now,move_candidate,path_risk,last_seen,note FROM directories ORDER BY CASE relevance WHEN '强相关' THEN 1 WHEN '上游相关' THEN 2 WHEN '弱相关' THEN 3 ELSE 4 END, path",
        "文件清单": "SELECT path,name,ext,category,stage,relevance,size_bytes FROM files ORDER BY relevance,category,path",
        "脚本职责": "SELECT path,project_id,canonical_stage_id,component_role,legacy_stage,responsibility,refs_external_paths FROM scripts ORDER BY project_id,canonical_stage_id,component_role,path",
        "模块注册表": "SELECT project_id,canonical_module,canonical_path,component_role,legacy_module,legacy_path,migration_status,public_api,test_evidence,notes FROM module_registry ORDER BY project_id,canonical_module",
        "外部数据依赖": "SELECT external_path,referenced_by_script,line_no,purpose,location FROM external_data_refs ORDER BY location,referenced_by_script",
        "外部数据盘": "SELECT path,location,description,referenced FROM external_disks",
        "流水线阶段": "SELECT stage,name,input,output,gate_status,evidence_dir FROM pipeline_stages ORDER BY CAST(stage AS REAL), stage",
        "时间运行批次": "SELECT run_id,kind,target_utc,parent_run_id,source_profile,status,manifest_path,note FROM time_runs ORDER BY run_id,source_profile",
        "元信息": "SELECT key,value FROM meta",
    }
    # 删除默认 sheet
    wb.remove(wb.active)
    for title, sql in sheets.items():
        ws = wb.create_sheet(title=title)
        cur = conn.execute(sql)
        cols = [d[0] for d in cur.description]
        ws.append(cols)
        for c in ws[1]:
            c.font = header_font
            c.fill = header_fill
            c.alignment = wrap
        for row in cur.fetchall():
            ws.append(list(row))
        # 列宽自适应（粗略）
        for i, col in enumerate(cols, 1):
            maxlen = max([len(str(col))] + [len(str(r[i-1])) if r[i-1] is not None else 0 for r in [cols]] )
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(max(maxlen * 1.6, 10), 60)
        ws.freeze_panes = "A2"
    conn.close()
    wb.save(XLSX_PATH)
    print(f"[OK] xlsx 导出: {XLSX_PATH}")


def fetch_dicts(conn: sqlite3.Connection, sql: str) -> list[dict[str, object]]:
    cur = conn.execute(sql)
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, row)) for row in cur.fetchall()]


def write_markdown_table(path: Path, rows: list[dict[str, object]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "| " + " | ".join(columns) + " |",
        "| " + " | ".join(["---"] * len(columns)) + " |",
    ]
    for row in rows:
        vals = []
        for col in columns:
            text = str(row.get(col, "") or "").replace("\n", " ").replace("|", "\\|")
            vals.append(text)
        lines.append("| " + " | ".join(vals) + " |")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def export_workspace_reports(db_path: Path = DB_PATH) -> None:
    WORKSPACE_DIR.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    dirs = fetch_dicts(
        conn,
        "SELECT path,relevance,role,file_count,size_text,referenced_by_code,exists_now,move_candidate,path_risk,note "
        "FROM directories ORDER BY CASE relevance WHEN '强相关' THEN 1 WHEN '上游相关' THEN 2 WHEN '弱相关' THEN 3 ELSE 4 END, path",
    )
    scripts = fetch_dicts(
        conn,
        "SELECT filename,project_id,canonical_stage_id,component_role,legacy_stage,responsibility,refs_external_paths FROM scripts ORDER BY project_id,canonical_stage_id,component_role,filename",
    )
    modules = fetch_dicts(
        conn,
        "SELECT project_id,canonical_module,canonical_path,component_role,legacy_module,legacy_path,migration_status,public_api,test_evidence,notes FROM module_registry ORDER BY project_id,canonical_module",
    )
    stages = fetch_dicts(
        conn,
        "SELECT stage,name,input,output,gate_status,evidence_dir FROM pipeline_stages ORDER BY id",
    )
    refs = fetch_dicts(
        conn,
        "SELECT external_path,referenced_by_script,line_no,purpose,location FROM external_data_refs ORDER BY location,referenced_by_script",
    )
    registry = fetch_dicts(
        conn,
        "SELECT project_id,canonical_stage_id,component_role,display_label,name,status,legacy_labels,evidence_paths,do_not_merge_with,notes FROM stage_registry ORDER BY project_id,stage_order,canonical_stage_id",
    )
    aliases = fetch_dicts(
        conn,
        "SELECT project_id,alias,canonical_stage_id,confidence,reason FROM stage_aliases ORDER BY project_id,alias",
    )
    artifacts = fetch_dicts(
        conn,
        "SELECT project_id,canonical_stage_id,component_role,artifact_type,path,legacy_stage_label,role,file_count,ext_summary,summary FROM artifact_index ORDER BY project_id,canonical_stage_id,artifact_type,path",
    )
    data_audits = fetch_dicts(
        conn,
        "SELECT project_id,audit_id,component_role,canonical_stage_id,related_stage_ids,data_domain,product_family,audit_scope,status,primary_path,output_root,notes FROM data_product_audits ORDER BY component_role,canonical_stage_id,audit_id",
    )
    violations = fetch_dicts(
        conn,
        "SELECT project_id,path,legacy_label,issue_type,severity,suggested_canonical_stage_id,suggested_new_path,reason FROM naming_violations ORDER BY severity DESC,project_id,path",
    )
    conn.close()

    repo_paths, _ = governance_check.all_repo_candidate_paths()
    absolute_path_warning_count = sum(
        "machine-local absolute path" in finding.message
        for finding in governance_check.check_paths(repo_paths, set(), baseline_mode=True)
    )
    dynamic_stage_loader_count = sum(
        "dynamic stage-script loading" in finding.message
        for finding in governance_check.check_dynamic_stage_loading(repo_paths, baseline_mode=False)
    )

    compact_artifacts = [
        row for row in artifacts
        if row.get("artifact_type") == "directory_summary"
        or row.get("role") in {"config", "evidence", "presentation_builder", "stage_report", "table"}
        or (
            row.get("role") in TIME_RUN_ARTIFACT_ROLES
            and row.get("artifact_type") in {"config_or_manifest", "report", "workbook"}
        )
    ]

    readme = f"""# GEO-ring Cloud Workspace

Generated: `{GENERATED_AT}`

This folder is a lightweight control surface for the GEO-ring Cloud project. It intentionally does not copy large data products.

## Source of Truth

- Main code: `{ROOT / "third_report" / "code" / "geo_ring_cloud_stage1"}`
- Stage1 products: `{ROOT / "geo_ring_cloud_stage1"}`
- Time-run products: `{ROOT / "geo_ring_cloud_stage1_time_runs"}`
- Evidence pack: `{ROOT / "geo_ring_cloud_stage1_evidence_pack"}`
- Index database: `{db_path}`

## Files Here

- `directory_classification.md`: relevance, existence, path risk, and archive candidacy.
- `architecture.md`: authoritative module boundaries and target physical structure.
- `engineering_status.md`: generated engineering-health snapshot and prioritized debt.
- `script_inventory.md`: current GEO-ring Cloud stage scripts and non-stage components.
- `module_registry.md`: canonical Python modules, compatibility shims, public APIs, and migration evidence.
- `pipeline_stages.md`: stage-level inputs, outputs, and evidence directories.
- `path_mapping.md`: code/data path dependencies and override strategy.
- `archive_manifest_dry_run.csv`: dry-run archive candidates generated before physical moves.
- `stage_registry.md`: canonical stage taxonomy and collision guards.
- `artifact_index.md`: compact project-memory view of directory summaries and high-value reports/manifests; query SQLite/XLSX for complete artifact rows.
- `data_product_audits.md`: horizontal index of generic and stage-scoped EO product inspections.
- `legacy_aliases.md`: legacy labels mapped to canonical stage IDs.
- `naming_policy.md`: naming rules for new work and known non-canonical labels.
- `engineering_policy.md`: enforceable engineering contract for humans and AI agents.
- Reproducible environment: `{ROOT / "third_report" / "code" / "geo_ring_cloud_stage1" / "environment.yml"}`
- Local/CI quality gate: `python _GEO_RING_CLOUD_INDEX\\ci_check.py`
"""
    (WORKSPACE_DIR / "README.md").write_text(readme, encoding="utf-8")
    architecture = """# GEO-ring Cloud Architecture

## 权威边界

- 控制面：`_GEO_RING_CLOUD_WORKSPACE` 与 `_GEO_RING_CLOUD_INDEX`，负责 taxonomy、检索、治理和项目记忆。
- 代码面：`third_report/code/geo_ring_cloud_stage1`，是 Geo Ring Cloud 主代码唯一权威入口。
- 数据面：`data`、`geo_ring_cloud_stage1`、`geo_ring_cloud_stage1_time_runs`、`data_check_report`、`geo_geometry_check`，保留原位并通过 `path_config.py` 配置。
- 证据面：stage manifest、关键 CSV/Markdown 索引和 evidence pack；大体量二进制产物不进入 Git。

## 逻辑模块

| layer | ownership | examples |
| --- | --- | --- |
| configuration | 路径、数据源 ID、环境覆盖、依赖契约 | `geo_ring_cloud.paths`, `geo_ring_cloud_path_configuration.ps1`, `geo_ring_cloud.pipeline_layout`, `geo_ring_cloud.sources`, `environment.yml` |
| lineage | manifest、commit、输入输出追踪 | `geo_ring_cloud.lineage` |
| adapters | 产品读取、格式适配、变量解码 | `geo_ring_cloud.adapters.claas3`, `geo_ring_cloud.adapters.epic`, `geo_data_audit/` |
| semantics | 云代码含义、display/fusion 有效性与质量规则 | `geo_ring_cloud.cloud_semantics` |
| reprojection | 原生定位、坐标规范化、最近邻重投影 | `geo_ring_cloud.reprojection` |
| geometry | GEO 轨道参数、球面/ECEF VZA | `geo_ring_cloud.geometry` |
| fusion support | 重投影目录、GEO 几何权重、候选源构建 | `geo_ring_cloud.fusion_support` |
| overlap metrics | 二值统计、源边界、重叠区 quicklook | `geo_ring_cloud.overlap` |
| audit semantics | 数据资产审计的可测试语义修正规则 | `geo_ring_cloud.data_asset_audit` |
| stage pipeline | 单一 canonical stage 的科学处理与验证 | `stage_09d_*`, `stage_10_*` |
| orchestration | 跨阶段实验、批处理、time-run matrix | `geo_ring_cloud_experiment_profile_pair.py`, `geo_ring_cloud_time_run_matrix.py` |
| diagnostics | 可复用指标、采样和分层统计 | `geo_ring_cloud.diagnostics.epic_pair` |
| presentation | 代表性图、组会材料生成 | `stage_10/stage_10_make_*` |
| tests | 轻量单元、smoke 与回归测试；生成物只放 `tests/_tmp` | `tests/` |

## 物理迁移原则

`geo_ring_cloud/` 是共享 Python API 的权威 package；顶层同名旧模块只允许作为 compatibility shim。当前已迁移路径配置、pipeline layout、云语义、重投影、GEO 几何、融合支撑、重叠统计、数据资产审计语义、数组摘要统计、数据源注册、lineage、run discovery、通用产品读取、quicklook、artifact IO、CLAAS-3/EPIC 产品适配器和 EPIC 配对诊断。`pipeline_support` 已降为纯兼容 facade，不得包含实现逻辑。其余扁平历史 stage 脚本不得为目录美观一次性移动；只有在导入引用、运行器路径、证据引用和 rollback manifest 均验证后，才分批迁移。

新 stage 若只有一个脚本，可使用 `stage_XX_<purpose>.py`；若有多个脚本，必须放入 `stage_XX_<purpose>/`。跨阶段工具不得伪造组合 stage，必须使用 `geo_ring_cloud_<role>_<purpose>.py`、声明 `COMPONENT_ROLE`，并在 manifest 中记录 `related_stage_ids`。
"""
    (WORKSPACE_DIR / "architecture.md").write_text(architecture, encoding="utf-8")

    warning_count = sum(1 for row in violations if row.get("severity") == "warning")
    geo_stage_count = sum(1 for row in registry if row.get("project_id") == PROJECT_ID)
    time_run_root_count = sum(
        1 for path in (ROOT / "geo_ring_cloud_stage1_time_runs").iterdir()
        if path.is_dir()
    ) if (ROOT / "geo_ring_cloud_stage1_time_runs").exists() else 0
    if absolute_path_warning_count:
        path_debt_status = "- 仍有机器本地绝对路径和非 canonical 命名；普通模式保留 warning，新增污染会被 hook 阻断。"
        path_debt_priority = "2. P1：逐批参数化仍活跃脚本中的机器本地绝对路径，并保持默认路径行为不变。"
    else:
        path_debt_status = "- 活跃项目代码中的机器本地绝对路径 warning 已清零；历史非 canonical 命名继续由 alias/baseline 吸收。"
        path_debt_priority = "2. P1：保持机器本地绝对路径 warning 为零，并为新增 Python/PowerShell 路径执行治理门禁。"
    if dynamic_stage_loader_count:
        dynamic_loader_status = f"- 剩余历史动态阶段加载：{dynamic_stage_loader_count} 个文件。"
        dynamic_loader_priority = f"3. P1：将剩余 {dynamic_stage_loader_count} 个动态加载点迁移到专责模块。"
    else:
        dynamic_loader_status = "- 阶段脚本之间的动态实现加载已清零；Stage 05/06/07 主链均使用静态 package API。"
        dynamic_loader_priority = "3. P1：保持动态阶段加载为零；复用逻辑必须进入已登记 package API。"
    engineering_status = f"""# GEO-ring Cloud Engineering Status

Generated: `{GENERATED_AT}`

## 当前规模

- 索引脚本：{len(scripts)}
- canonical shared modules：{len(modules)}
- canonical stages：{geo_stage_count}
- SQLite 详细 artifact 记录：{len(artifacts)}
- Markdown 快查 artifact 记录：{len(compact_artifacts)}
- data product audits：{len(data_audits)}
- time-run 顶层目录：{time_run_root_count}
- 已登记历史命名 warning：{warning_count}
- 历史绝对路径 warning 文件：{absolute_path_warning_count}
- 历史动态阶段加载 warning 文件：{dynamic_stage_loader_count}

## 已建立的工程能力

- Git 仓库、远端、`.gitignore`、`.gitattributes` 与本地 pre-commit hook。
- canonical stage taxonomy、artifact index、data product audit index 和跨项目 collision guard。
- Python `geo_ring_cloud.paths` 与 PowerShell `geo_ring_cloud_path_configuration.ps1` 共享环境变量契约；统一 lineage manifest helper 与 staged governance check。
- `geo_ring_cloud` package、`pyproject.toml`、module registry 与旧 import compatibility shims。
- 已验证直接依赖基线、统一 `ci_check.py` 入口与 GitHub 轻量 CI 门禁。
- 大数据、time-run、图片、Office 文件和生成数据库默认不进入 Git。

## 尚未达到的目标

- `stage1_common.py` 已降为 compatibility shim；`pipeline_support` 已降为纯兼容 facade，layout、cloud semantics、重投影、GEO 几何、融合支撑、重叠统计、数据资产审计语义、产品读取、quicklook、artifact IO 与数组摘要统计均已拆入专责模块。
{dynamic_loader_status}
{path_debt_status}
- `environment.yml` 已固定已验证的直接依赖；跨平台传递依赖锁仍应在正式实验发布时按平台生成。
- 一部分旧 time-run 使用 `stage0910` 等组合标签；为保障续跑暂保留，只作为 legacy alias，不得用于新组件命名。

## 优先级

1. P0：任何新增 governance error 必须在提交前清零。
{path_debt_priority}
{dynamic_loader_priority}
4. P2：为正式实验发布生成平台化传递依赖锁；大数据集成测试继续本地运行。
5. P2：按依赖审计结果渐进迁移扁平脚本，禁止一次性大搬迁。
"""
    (WORKSPACE_DIR / "engineering_status.md").write_text(engineering_status, encoding="utf-8")
    write_markdown_table(
        WORKSPACE_DIR / "directory_classification.md",
        dirs,
        ["path", "relevance", "role", "file_count", "size_text", "referenced_by_code", "exists_now", "move_candidate", "path_risk", "note"],
    )
    write_markdown_table(
        WORKSPACE_DIR / "script_inventory.md",
        scripts,
        ["filename", "project_id", "canonical_stage_id", "component_role", "legacy_stage", "responsibility", "refs_external_paths"],
    )
    write_markdown_table(
        WORKSPACE_DIR / "module_registry.md",
        modules,
        ["project_id", "canonical_module", "canonical_path", "component_role", "legacy_module", "legacy_path", "migration_status", "public_api", "test_evidence", "notes"],
    )
    write_markdown_table(
        WORKSPACE_DIR / "pipeline_stages.md",
        stages,
        ["stage", "name", "input", "output", "gate_status", "evidence_dir"],
    )
    write_markdown_table(
        WORKSPACE_DIR / "path_mapping.md",
        refs,
        ["external_path", "referenced_by_script", "line_no", "purpose", "location"],
    )
    write_markdown_table(
        WORKSPACE_DIR / "stage_registry.md",
        registry,
        ["project_id", "canonical_stage_id", "component_role", "display_label", "name", "status", "legacy_labels", "evidence_paths", "do_not_merge_with", "notes"],
    )
    write_markdown_table(
        WORKSPACE_DIR / "artifact_index.md",
        compact_artifacts,
        ["project_id", "canonical_stage_id", "component_role", "artifact_type", "path", "legacy_stage_label", "role", "file_count", "ext_summary", "summary"],
    )
    write_markdown_table(
        WORKSPACE_DIR / "data_product_audits.md",
        data_audits,
        ["project_id", "audit_id", "component_role", "canonical_stage_id", "related_stage_ids", "data_domain", "product_family", "audit_scope", "status", "primary_path", "output_root", "notes"],
    )
    write_markdown_table(
        WORKSPACE_DIR / "legacy_aliases.md",
        aliases,
        ["project_id", "alias", "canonical_stage_id", "confidence", "reason"],
    )
    write_markdown_table(
        WORKSPACE_DIR / "naming_violations.md",
        violations,
        ["project_id", "path", "legacy_label", "issue_type", "severity", "suggested_canonical_stage_id", "suggested_new_path", "reason"],
    )
    engineering_policy = """# GEO-ring Cloud Engineering Policy

This document is the enforceable engineering contract for Geo Ring Cloud work.
It applies to humans and AI agents.

## Required workflow

- MUST check `architecture.md`, `engineering_status.md`, `module_registry.md`, `stage_registry.md`, `artifact_index.md`, `data_product_audits.md`, and the SQLite index before creating new code or reports.
- MUST reuse existing scripts, manifests, reports, and products when they already answer the task.
- MUST decide the `project_id + canonical_stage_id` before naming files.
- MUST run `python _GEO_RING_CLOUD_INDEX\\build_index.py` after adding or changing stage scripts.
- Existing-stage refactors MUST stage refreshed `artifact_index.md` when artifact semantics change; otherwise refreshed `engineering_status.md` is acceptable. New stages MUST stage the full stage/artifact/audit index set.
- MUST run `python _GEO_RING_CLOUD_INDEX\\governance_check.py --staged` before commit.
- MUST use the checked-in `environment.yml` as the default scientific dependency baseline and run `python _GEO_RING_CLOUD_INDEX\\ci_check.py --scientific-tests` for core-code changes.

## Naming and identity

- MUST use canonical stage IDs for new stage-owned files, such as `stage_10p2_approx_fov_report.md`.
- MUST NOT create new `Step*`, `stage10*`, `Stage10*`, or `10_stage*` names.
- MUST use `geo_ring_cloud_<role>_<purpose>.py` for new non-stage core utilities.
- MUST place reusable shared APIs in the `geo_ring_cloud` package and import them through their canonical module names.
- Package adapters and diagnostics MUST NOT import or dynamically load stage scripts; dependencies flow from stages to shared APIs.
- Stage scripts MUST NOT dynamically load one another to reuse implementation; extract shared logic into a registered `geo_ring_cloud.*` module and use a normal import. Registered historical loaders are migration warnings only.
- `geo_ring_cloud.pipeline_support` is a transitional compatibility facade. It MUST contain only imports, export metadata, and aliases; active stage/component code MUST NOT import it, and new shared responsibilities MUST use focused package modules.
- Staged code MUST NOT import registered top-level compatibility shims; use canonical `geo_ring_cloud.*` modules.
- Only the dedicated compatibility boundary test may import legacy shims, through the governance allowlist.
- MUST NOT add implementation logic to top-level compatibility shims recorded in `module_registry.md`.
- MUST NOT treat `geo_ring_cloud.stage_09` and `epic_ceres.stage_09` as the same stage.

## Output lineage

- New stage outputs MUST include a manifest with `project_id`, `canonical_stage_id`, generating script, inputs, outputs, parameters, timestamp, and commit when available.
- Non-stage run manifests MUST include `component_role` and `related_stage_ids`; they MUST NOT place a component label in `canonical_stage_id`.
- Reports SHOULD be Chinese-first, with English retained for technical terms and variable names.
- Key outputs SHOULD include concise CSV/Markdown indexes instead of relying only on directory names.
- Generic data/product inspections SHOULD be indexed in `data_product_audits.md`; stage-scoped inspections should keep `related_stage_ids`.

## Path and artifact rules

- Python code MUST use `geo_ring_cloud.paths`; PowerShell orchestration MUST dot-source `geo_ring_cloud_path_configuration.ps1` or use the same `GEO_RING_*` environment-variable contract.
- Active project code MUST NOT hard-code any machine-local drive path unless it is one of the two explicitly allowlisted canonical path-configuration files.
- Core code MUST NOT depend on `_NON_GEO_ARCHIVE`, `second_report`, `forth`, or EPIC-CERES code/output paths.
- Raw data, time runs, evidence packs, SQLite/XLSX indexes, PPTX, images, NetCDF/HDF/HDF5, NPZ, and other large generated artifacts MUST stay out of Git by default.
- GitHub CI MUST remain independent of local large-data paths; real-data integration tests are explicit local checks.

## Enforcement levels

- New violations are errors in the staged governance check.
- Historical naming and path debt remains warnings unless `--strict` is used.
- Historical warnings should be cleaned in dedicated cleanup work, not opportunistically mixed into scientific changes.
"""
    (WORKSPACE_DIR / "engineering_policy.md").write_text(engineering_policy, encoding="utf-8")

    naming_policy = """# GEO-ring Cloud Naming Policy

## Canonical identifiers

- Use `project_id + canonical_stage_id` for every stage decision.
- Main project namespace: `geo_ring_cloud`.
- Stage IDs use lowercase ASCII: `stage_00`, `stage_03_5`, `stage_06c`, `stage_07p_b`, `stage_07v2`, `stage_09d`, `stage_10p2`.
- Do not use `Step` for project-level phases. `Step` may only describe an internal procedure inside a script or report.

## New file and directory names

- Prefix new stage-owned files with the canonical stage ID, for example `stage_09d_full_pixel_diagnostics_report.md`.
- New stage-owned directories must also use the canonical stage ID, for example `stage_10p2_approx_fov_aggregation`.
- Put substep numbers after the stage directory or in report sections, for example `stage_09d/00_sample_manifest`.
- Reusable shared APIs belong in the `geo_ring_cloud` package, use lowercase `snake_case.py`, declare `COMPONENT_ROLE`, and must be registered in `module_registry.md`.
- Executable non-stage utilities at the code root must use `geo_ring_cloud_<role>_<purpose>.py`, declare `COMPONENT_ROLE`, and avoid fake or combined stages. Roles include `runner`, `experiment_runner`, `downloader`, `evidence_pack_builder`, `summary_helper`, and `presentation_builder`.
- Top-level compatibility shims may retain historical names only when registered; they must contain imports and metadata, not implementation logic.
- Generic EO data/product inspections must use `component_role=data_product_audit`; keep legacy `third_report/code/geo_data_audit` paths until references are audited, and index them in `data_product_audits.md`.
- Do not create new `Step*`, `stage10*`, `Stage10*`, `09_stage*`, or numeric-prefix stage names.

## Collision rules

- `geo_ring_cloud.stage_09` is not `epic_ceres.stage_09`.
- `geo_ring_cloud.stage_09` is not `epic_ceres.stage_09_5`.
- `research_tracker` labels such as `Step9`, `BuildStep9`, and `Stage9` are untrusted legacy inference labels until reviewed in `stage_registry`.

## Migration rule

Historical files are not renamed by default. Rename only after code references, evidence references, and a rollback manifest are checked.

## Enforcement rule

- Newly added non-canonical stage names are errors.
- Existing historical names remain warnings during normal checks.
- Use `python _GEO_RING_CLOUD_INDEX\\governance_check.py --all --strict` for strict audit mode.
"""
    (WORKSPACE_DIR / "naming_policy.md").write_text(naming_policy, encoding="utf-8")

    suggestions = f"""# GEO-ring Cloud 整理建议

> 生成时间：`{GENERATED_AT}`

## 结论

- 强相关目录保持原位，避免破坏现有流水线和历史产物路径。
- 无关目录只在归档 manifest 确认零引用后移动到 `{ARCHIVE_DIR}`。
- 运行时代码路径通过 `geo_ring_cloud.paths` 参数化；`path_config.py` 仅保留旧 import 兼容。

## 高优先级保留

- `geo_ring_cloud_stage1`
- `geo_ring_cloud_stage1_time_runs`
- `geo_ring_cloud_stage1_evidence_pack`
- `third_report/code/geo_ring_cloud_stage1`
- `data_check_report`
- `geo_geometry_check`
- `data`

## 可归档候选

见 `_NON_GEO_ARCHIVE/_move_manifest.csv` 或本 workspace 的 `archive_manifest_dry_run.csv`。

## 命名体系

- 权威阶段表见 `_GEO_RING_CLOUD_WORKSPACE/stage_registry.md`。
- 旧名称映射见 `_GEO_RING_CLOUD_WORKSPACE/legacy_aliases.md`。
- 关键产物索引见 `_GEO_RING_CLOUD_WORKSPACE/artifact_index.md`。
- 共享模块与旧 import 映射见 `_GEO_RING_CLOUD_WORKSPACE/module_registry.md`。
- 命名规则见 `_GEO_RING_CLOUD_WORKSPACE/naming_policy.md`。
"""
    (OUT_DIR / "整理建议.md").write_text(suggestions, encoding="utf-8")
    print(f"[OK] workspace reports: {WORKSPACE_DIR}")
    print(f"[OK] 整理建议: {OUT_DIR / '整理建议.md'}")


if __name__ == "__main__":
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    built_db = build_sqlite()
    export_xlsx(built_db)
    export_workspace_reports(built_db)
    print("[DONE]")
